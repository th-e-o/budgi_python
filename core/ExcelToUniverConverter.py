from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass, asdict
from datetime import datetime
from typing import Dict, Any, List, Tuple, Optional, Set
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

from modules.excel.excel_utils import ExcelUtils


class ExcelToUniverConverter:
    """Convert Excel workbook to Univer Sheets IWorkbookData format."""

    def __init__(self, workbook: Workbook):
        self.workbook = workbook
        self.theme_colors = ExcelUtils.get_theme_colors(workbook)
        self.style_registry: Dict[str, Dict[str, Any]] = {}
        self.style_hash_to_id: Dict[str, str] = {}
        self.style_counter = 0

        self.DEFAULT_FONT = Font()
        self.DEFAULT_PATTERN_FILL = PatternFill()
        self.DEFAULT_BORDER = Border()
        self.DEFAULT_ALIGNMENT = Alignment()

        self.style_map = {
            'thin': 1,
            'hair': 2,
            'medium': 8,
            'thick': 13,
            'dashed': 4,
            'dotted': 3,
            'double': 7
        }

        self.border_map = {
            'top': 't',
            'bottom': 'b',
            'left': 'l',
            'right': 'r'
        }

    def convert(self) -> Dict[str, Any]:
        """Main conversion method that returns IWorkbookData."""
        workbook_data = {
            "id": self._generate_id(),
            "name": self._get_workbook_name(),
            "locale": "FR_FR",
            "styles": {},
            "sheets": {},
            "sheetOrder": [],
            "definedNames": [{'n': p, 'formulaRefOrString': n.attr_text.replace('$','')} for p, n in self.workbook.defined_names.items()]
        }

        # Process each sheet
        for sheet in self.workbook.worksheets:
            sheet_id = self._generate_sheet_id(sheet.title)
            workbook_data["sheetOrder"].append(sheet_id)
            workbook_data["sheets"][sheet_id] = self._convert_sheet(sheet, sheet_id)

        # Add collected styles
        workbook_data["styles"] = self.style_registry

        return workbook_data

    def _generate_id(self) -> str:
        """Generate a unique ID for the workbook."""
        return hashlib.md5(f"workbook_{datetime.now().isoformat()}".encode()).hexdigest()[:6]

    def _generate_sheet_id(self, sheet_name: str) -> str:
        """Generate a unique ID for a sheet."""
        return hashlib.md5(f"sheet_{sheet_name}_{datetime.now().isoformat()}".encode()).hexdigest()[:20]

    def _get_workbook_name(self) -> str:
        """Get workbook name or default."""
        # Excel doesn't store workbook name in the same way, so we'll use a default
        return "Imported Workbook"

    def _convert_sheet(self, sheet: Worksheet, sheet_id: str) -> Dict[str, Any]:
        """Convert a single worksheet to IWorksheetData format."""
        max_row, max_col = ExcelUtils.get_data_only_range(sheet)

        sheet_data = {
            "id": sheet_id,
            "name": sheet.title,
            "tabColor": self._get_tab_color(sheet),
            "hidden": 0 if sheet.sheet_state == 'visible' else 1,
            "rowCount": max_row + 1,
            "columnCount": max_col + 1,
            "zoomRatio": sheet.sheet_view.zoomScale / 100 if sheet.sheet_view.zoomScale else 1,
            "freeze": self._get_freeze_panes(sheet),
            "scrollTop": 0,
            "scrollLeft": 0,
            #"defaultColumnWidth": 73,
            #"defaultRowHeight": 23,
            "mergeData": self._get_merge_data(sheet),
            "cellData": self._get_cell_data(sheet, max_row, max_col),
            "rowData": self._get_row_data(sheet, max_row),
            "columnData": self._get_column_data(sheet, max_col),
            "showGridlines": 1 if sheet.sheet_view.showGridLines else 0,
            "rowHeader": {
                "width": 46,
                "hidden": 0
            },
            "columnHeader": {
                "height": 20,
                "hidden": 0
            },
            "rightToLeft": 0
        }

        return sheet_data

    def _get_tab_color(self, sheet: Worksheet) -> str:
        """Extract tab color from sheet."""
        if sheet.sheet_properties.tabColor:
            color = sheet.sheet_properties.tabColor
            if hasattr(color, 'rgb'):
                return f"#{color.rgb[-6:]}" if color.rgb else ""
            elif hasattr(color, 'theme') and hasattr(color, 'tint'):
                return f"#{ExcelUtils.theme_and_tint_to_rgb(self.theme_colors, color.theme, color.tint or 0)}"
        return ""

    def _get_freeze_panes(self, sheet: Worksheet) -> Dict[str, int]:
        """Extract freeze pane settings."""
        freeze = {
            "startRow": -1,
            "startColumn": -1,
            "ySplit": 0,
            "xSplit": 0
        }

        if sheet.freeze_panes:
            col_letter, row = coordinate_from_string(sheet.freeze_panes)
            col = column_index_from_string(col_letter)
            freeze["startRow"] = row - 1
            freeze["startColumn"] = col - 1
            freeze["ySplit"] = row - 1
            freeze["xSplit"] = col - 1

        return freeze

    def _get_merge_data(self, sheet: Worksheet) -> List[Dict[str, int]]:
        """Extract merged cell ranges."""
        merge_data = []
        for merge_range in sheet.merged_cells.ranges:
            merge_data.append({
                "startRow": merge_range.min_row - 1,
                "endRow": merge_range.max_row - 1,
                "startColumn": merge_range.min_col - 1,
                "endColumn": merge_range.max_col - 1
            })
        return merge_data

    def _get_cell_data(self, sheet: Worksheet, max_row: int, max_col: int) -> Dict[str, Dict[str, Dict[str, Any]]]:
        """Extract cell data with optimized style references."""
        cell_data = {}

        for row_of_cells in sheet.iter_rows(max_row=max_row, max_col=max_col):
            row_data = {}
            for cell in row_of_cells:
                cell_info = self._convert_cell(cell)
                if cell_info:
                    # cell.column is 1-based, we need 0-based
                    row_data[str(cell.column - 1)] = cell_info

            if row_data:
                # All cells in row_of_cells have the same row index
                cell_data[str(cell.row - 1)] = row_data

        return cell_data

    def _convert_cell(self, cell: Cell) -> Optional[Dict[str, Any]]:
        """Convert a single cell to Univer format."""
        # Skip empty cells with no formatting
        if cell.value is None and self._is_default_cell_style(cell):
            return None

        cell_data = {}

        # Handle cell value
        if cell.value is not None:
            if cell.data_type != 'f':
                cell_data["v"] = self._convert_cell_value(cell)

            # Set cell type if needed
            if cell.data_type == 'f':
                if not isinstance(cell.value, (str, int, float, bool, type(None))):
                    cell_value = cell.value.text if hasattr(cell.value, 'text') else str(cell.value)
                else:
                    cell_value = str(cell.value)
                cell_data["f"] = cell_value
            elif cell.data_type == 'n' and isinstance(cell.value, (int, float)):
                cell_data["t"] = 2  # Number type
            elif cell.data_type == 's':
                cell_data["t"] = 1  # String type

        # Handle cell style
        style_id = self._get_or_create_style_id(cell)
        if style_id:
            cell_data["s"] = style_id

        return cell_data if cell_data else None

    def _convert_cell_value(self, cell: Cell) -> Any:
        """Convert cell value to appropriate format."""
        value = cell.value

        if isinstance(value, datetime):
            # Convert datetime to Excel serial number
            return (value - datetime(1899, 12, 30)).total_seconds() / 86400
        elif isinstance(value, bool):
            return 1 if value else 0
        elif value is None:
            return ""
        else:
            return value

    def _is_default_cell_style(self, cell: Cell) -> bool:
        """Check if cell has default styling."""
        return (
                cell.font == self.DEFAULT_FONT and
                cell.fill == self.DEFAULT_PATTERN_FILL and
                cell.border == self.DEFAULT_BORDER and
                cell.alignment == self.DEFAULT_ALIGNMENT and
                cell.number_format == 'General'
        )

    def _get_or_create_style_id(self, cell: Cell) -> Optional[str]:
        """
        Get existing style ID or create a new one for the cell's style.
        This version is optimized to avoid repeated hashing and looping.
        """
        style_data = self._extract_cell_style(cell)

        # If the cell has no special styling, return None.
        if not style_data:
            return None

        style_hash = hashlib.md5(json.dumps(style_data, sort_keys=True).encode()).hexdigest()

        if style_hash in self.style_hash_to_id:
            return self.style_hash_to_id[style_hash]

        # Style not in cache
        style_id = f"s{self.style_counter}"
        self.style_counter += 1

        # Store the style in the registry and cache
        self.style_registry[style_id] = style_data
        self.style_hash_to_id[style_hash] = style_id

        return style_id


    def _extract_cell_style(self, cell: Cell) -> Optional[Dict[str, Any]]:
        """Extract style data from cell in Univer format."""
        style = {}

        font = cell.font
        alignment = cell.alignment
        number_format = cell.number_format
        fill = cell.fill  # For background color
        border = cell.border  # For border extraction

        if font != self.DEFAULT_FONT:
            if font.name and font.name != 'Calibri':
                style["ff"] = font.name
            if font.sz and font.sz != 11:
                style["fs"] = font.sz
            if font.bold:
                style["bl"] = 1
            if font.italic:
                style["it"] = 1
            if font.underline:
                style["ul"] = {"s": 1}
            if font.strike:
                style["st"] = {"s": 1}

            font_color = self._extract_color(font.color)
            if font_color and font_color != "000000":
                style["cl"] = {"rgb": f"#{font_color}"}

        # Background color
        bg_color, _ = ExcelUtils.extract_cell_colors(fill, font, self.theme_colors)
        if bg_color and bg_color != "FFFFFF":
            style["bg"] = {"rgb": f"#{bg_color}"}

        # Borders
        border_data = self._extract_borders(border)
        if border_data:
            style["bd"] = border_data

        # Alignment
        if alignment != self.DEFAULT_ALIGNMENT:
            if alignment.horizontal:
                ha_map = {"left": 1, "center": 2, "right": 3, "justify": 4}
                if alignment.horizontal in ha_map:
                    style["ht"] = ha_map[alignment.horizontal]
            if alignment.vertical:
                va_map = {"top": 1, "center": 2, "bottom": 3}
                if alignment.vertical in va_map:
                    style["vt"] = va_map[alignment.vertical]
            if alignment.wrap_text:
                style["tb"] = 3
            if alignment.text_rotation and alignment.text_rotation != 0:
                style["tr"] = {"a": alignment.text_rotation, "v": 0}

        # Number format
        if number_format and number_format != 'General':
            style["n"] = {"pattern": number_format}

        return style if style else None

    def _extract_color(self, color) -> Optional[str]:
        """Extract color value from openpyxl color object."""
        if color is None:
            return None

        if color.type == "rgb" and color.rgb:
            if color.rgb == '00000000':
                return 'FFFFFF' # Special case for transparent fill'
            else:
                return color.rgb[-6:]  # Return RGB without alpha channel
        elif color.type == 'theme':
            return ExcelUtils.theme_and_tint_to_rgb(self.theme_colors, color.theme, color.tint or 0)

        return None

    def _extract_borders(self, border) -> Optional[Dict[str, Any]]:
        """Extract border data in Univer format."""
        if border == self.DEFAULT_BORDER:
            return None

        border_data = {}

        for excel_side, univer_side in self.border_map.items():
            side = getattr(border, excel_side)
            if side and side.style:
                border_info = {
                    "s": self._convert_border_style(side.style),
                    "cl": {"rgb": f"#{self._extract_color(side.color) or '000000'}"}
                }
                border_data[univer_side] = border_info

        return border_data if border_data else None

    def _convert_border_style(self, style: str) -> int:
        """Convert Excel border style to Univer format."""
        # Map Excel border styles to Univer border style codes

        assert style in self.style_map, f"Unsupported border style: {style}"
        return self.style_map.get(style)

    def _get_row_data(self, sheet: Worksheet, max_row: int) -> Dict[str, Dict[str, Any]]:
        """Extract row data (heights, hidden status)."""
        row_data = {}

        for row_num in range(1, max_row + 1):
            row_dim = sheet.row_dimensions[row_num]
            if row_dim.customHeight or row_dim.hidden:
                row_info = {}
                if row_dim.customHeight and row_dim.ht:
                    row_info["h"] = row_dim.ht * 1.4
                if row_dim.hidden:
                    row_info["hd"] = 1

                if row_info:
                    row_data[str(row_num - 1)] = row_info

        return row_data

    def _get_column_data(self, sheet: Worksheet, max_col: int) -> Dict[str, Dict[str, Any]]:
        """Extract column data (widths, hidden status)."""
        column_data = {}

        for col_num in range(1, max_col + 1):
            col_letter = get_column_letter(col_num)
            if col_letter in sheet.column_dimensions:
                col_dim = sheet.column_dimensions[col_letter]
                col_info = {}

                if col_dim.width and col_dim.width != sheet.sheet_format.defaultColWidth:
                    # Convert Excel width to pixels (approximate)
                    col_info["w"] = int(col_dim.width * 7.5)
                if col_dim.hidden:
                    col_info["hd"] = 1

                if col_info:
                    column_data[str(col_num - 1)] = col_info

        return column_data