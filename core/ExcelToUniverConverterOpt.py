from __future__ import annotations

import hashlib
import json
from datetime import datetime
from typing import Dict, Any, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Border
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet

from modules.excel.excel_utils import ExcelUtils


class ExcelToUniverConverterOpt:
    """
    Correctly and efficiently converts an Excel workbook to Univer format
    by processing each unique style only once.
    """

    def __init__(self, workbook: Workbook):
        self.workbook = workbook
        self.theme_colors = ExcelUtils.get_theme_colors(workbook)

        # The final Univer style definitions
        self.style_registry: Dict[str, Dict[str, Any]] = {}

        # A cache to avoid re-calculating styles
        self.style_hash_to_id: Dict[str, str] = {}

        # The main lookup cache
        self.style_id_lookup: Dict[int, Optional[str]] = {}
        self.style_counter = 0

        self.style_map = {
            'thin': 1, 'hair': 2, 'medium': 8, 'thick': 13,
            'dashed': 4, 'dotted': 3, 'double': 7
        }
        self.border_map = {'top': 't', 'bottom': 'b', 'left': 'l', 'right': 'r'}

    def convert(self) -> Dict[str, Any]:
        """Main conversion method that returns IWorkbookData."""
        workbook_data = {
            "id": self._generate_id(),
            "name": "Imported Workbook",
            "locale": "FR_FR",
            "styles": {},
            "sheets": {},
            "sheetOrder": [],
            "definedNames": [{'n': p, 'formulaRefOrString': n.attr_text.replace('$', '')}
                             for p, n in self.workbook.defined_names.items()]
        }
        for sheet in self.workbook.worksheets:
            sheet_id = self._generate_sheet_id(sheet.title)
            workbook_data["sheetOrder"].append(sheet_id)
            workbook_data["sheets"][sheet_id] = self.convert_sheet(sheet, sheet_id=sheet_id)

        workbook_data["styles"] = self.style_registry
        return workbook_data

    def _get_or_create_style_id_for_cell(self, cell: Cell) -> Optional[str]:
        """
        The core of the optimization. Gets a cached Univer style ID for a cell's
        style_id, or computes and caches it if seen for the first time.
        """
        if cell.style_id in self.style_id_lookup:
            return self.style_id_lookup[cell.style_id]

        style_data = self._extract_style_from_cell(cell)

        if not style_data:
            self.style_id_lookup[cell.style_id] = None
            return None

        style_hash = hashlib.md5(json.dumps(style_data, sort_keys=True).encode()).hexdigest()

        if style_hash in self.style_hash_to_id:
            univer_style_id = self.style_hash_to_id[style_hash]
        else:
            univer_style_id = f"s{self.style_counter}"
            self.style_counter += 1
            self.style_registry[univer_style_id] = style_data
            self.style_hash_to_id[style_hash] = univer_style_id

        self.style_id_lookup[cell.style_id] = univer_style_id
        return univer_style_id

    def _extract_style_from_cell(self, cell: Cell) -> Dict[str, Any]:
        """
        Extracts the full, final style definition from a SINGLE cell object.
        This is the logic from your original, working-but-slow version.
        """
        style = {}
        font = cell.font
        fill = cell.fill
        border = cell.border
        alignment = cell.alignment
        number_format = cell.number_format

        # Font
        if font.name and font.name != 'Calibri': style["ff"] = font.name
        if font.sz and font.sz != 11: style["fs"] = float(font.sz)
        if font.bold: style["bl"] = 1
        if font.italic: style["it"] = 1
        if font.underline and font.underline != 'none': style["ul"] = {"s": 1}
        if font.strike: style["st"] = {"s": 1}
        font_color = self._extract_color(font.color)
        if font_color and font_color != "000000": style["cl"] = {"rgb": f"#{font_color}"}

        # Fill (Background)
        bg_color, _ = ExcelUtils.extract_cell_colors(fill, font, self.theme_colors)
        if bg_color and bg_color != "FFFFFF": style["bg"] = {"rgb": f"#{bg_color}"}

        # Border
        if border_data := self._extract_borders(border): style["bd"] = border_data

        # Alignment
        if alignment.horizontal:
            ha_map = {"left": 1, "center": 2, "right": 3, "justify": 4}
            if alignment.horizontal in ha_map: style["ht"] = ha_map[alignment.horizontal]
        if alignment.vertical:
            va_map = {"top": 1, "center": 2, "bottom": 3}
            if alignment.vertical in va_map: style["vt"] = va_map[alignment.vertical]
        if alignment.wrap_text: style["tb"] = 3
        if alignment.text_rotation and alignment.text_rotation != 0:
            style["tr"] = {"a": alignment.text_rotation, "v": 0}

        # Number Format
        if number_format and number_format != 'General':
            style["n"] = {"pattern": number_format}

        return style

    def _convert_cell(self, cell: Cell) -> Optional[Dict[str, Any]]:
        """Converts a single cell to Univer format using the cached style lookup."""
        univer_style_id = self._get_or_create_style_id_for_cell(cell)

        if cell.value is None and not univer_style_id:
            return None

        cell_data = {}
        if cell.value is not None:
            if cell.data_type != 'f':
                cell_data["v"] = self._convert_cell_value(cell.value)
            if cell.data_type == 'f':
                cell_value = cell.value if isinstance(cell.value, str) else str(cell.value)
                cell_data["f"] = f"={cell_value.lstrip('=')}"
            elif cell.data_type == 'n' and isinstance(cell.value, (int, float)):
                cell_data["t"] = 2
            elif cell.data_type == 's':
                cell_data["t"] = 1

        if univer_style_id:
            cell_data["s"] = univer_style_id

        return cell_data if cell_data else None

    def import_from_dataframe(self, df: Any, sheet_name: str,
                              start_row: int = 1, start_col: int = 1,
                              override_formula: bool = False) -> Dict[str, Any]:
        """
        Import data from a pandas DataFrame into an existing sheet and return the converted sheet data.

        :param df: The DataFrame to import.
        :param sheet_name: Name of the sheet to import the DataFrame into.
        :param start_row: Starting row (1-based) to place the DataFrame data.
        :param start_col: Starting column (1-based) to place the DataFrame data.
        :param override_formula: If True, will override existing formulas with DataFrame values.
        :return: Dictionary containing the converted sheet data in Univer format.
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")

        worksheet = self.workbook[sheet_name]

        # Calculate the data range
        num_rows, num_cols = df.shape

        # Write DataFrame values to the worksheet
        for row_idx in range(num_rows):
            for col_idx in range(num_cols):
                cell = worksheet.cell(row=start_row + row_idx, column=start_col + col_idx)

                # Get the DataFrame value
                df_value = df.iloc[row_idx, col_idx]

                # Skip NaN values
                if pd.isna(df_value):
                    continue

                # Only override if not a formula or if override_formula is True
                if cell.data_type == 'f' and not override_formula:
                    continue

                # Handle different data types
                if isinstance(df_value, str) and df_value.startswith('='):
                    # It's a formula
                    cell.value = df_value
                else:
                    # Regular value
                    cell.value = df_value

        # Convert the modified sheet using the existing convert_sheet method
        return self.convert_sheet(worksheet)

    def convert_sheet(self, sheet: Worksheet, sheet_id: str = None) -> Dict[str, Any]:
        max_row, max_col = ExcelUtils.get_data_only_range(sheet)
        if sheet_id is None:
            sheet_id = self._generate_sheet_id(sheet.title)
        return {
            "id": sheet_id,
            "name": sheet.title,
            "tabColor": self._get_tab_color(sheet),
            "hidden": 0 if sheet.sheet_state == 'visible' else 1,
            "rowCount": max_row + 1, "columnCount": max_col + 1,
            "zoomRatio": sheet.sheet_view.zoomScale / 100 if sheet.sheet_view.zoomScale else 1,
            "freeze": self._get_freeze_panes(sheet),
            "scrollTop": 0, "scrollLeft": 0,
            "mergeData": self._get_merge_data(sheet),
            "cellData": self._get_cell_data(sheet, max_row, max_col),
            "rowData": self._get_row_data(sheet, max_row),
            "columnData": self._get_column_data(sheet, max_col),
            "showGridlines": 1 if sheet.sheet_view.showGridLines else 0,
            "rowHeader": {"width": 46, "hidden": 0},
            "columnHeader": {"height": 20, "hidden": 0},
            "rightToLeft": 0
        }

    def _get_cell_data(self, sheet: Worksheet, max_row: int, max_col: int) -> Dict[str, Dict[str, Dict[str, Any]]]:
        cell_data = {}
        for row_idx in range(1, max_row + 1):
            row_data = {}
            for col_idx in range(1, max_col + 1):
                # Access cell by coordinate to ensure it exists for style processing
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell_info := self._convert_cell(cell):
                    row_data[str(col_idx - 1)] = cell_info
            if row_data:
                cell_data[str(row_idx - 1)] = row_data
        return cell_data

    def _generate_id(self) -> str:
        return hashlib.md5(f"workbook_{datetime.now().isoformat()}".encode()).hexdigest()[:6]

    def _generate_sheet_id(self, sheet_name: str) -> str:
        return hashlib.md5(f"sheet_{sheet_name}_{datetime.now().isoformat()}".encode()).hexdigest()[:20]

    def _get_tab_color(self, sheet: Worksheet) -> str:
        if sheet.sheet_properties.tabColor:
            color = sheet.sheet_properties.tabColor
            if hasattr(color, 'rgb'):
                return f"#{color.rgb[-6:]}" if color.rgb else ""
            elif hasattr(color, 'theme') and hasattr(color, 'tint'):
                return f"#{ExcelUtils.theme_and_tint_to_rgb(self.theme_colors, color.theme, color.tint or 0)}"
        return ""

    def _get_freeze_panes(self, sheet: Worksheet) -> Dict[str, int]:
        freeze = {"startRow": -1, "startColumn": -1, "ySplit": 0, "xSplit": 0}
        if sheet.freeze_panes:
            col_letter, row = coordinate_from_string(sheet.freeze_panes)
            col = column_index_from_string(col_letter)
            freeze.update({"startRow": row - 1, "startColumn": col - 1, "ySplit": row - 1, "xSplit": col - 1})
        return freeze

    def _get_merge_data(self, sheet: Worksheet) -> List[Dict[str, int]]:
        return [{
            "startRow": r.min_row - 1, "endRow": r.max_row - 1,
            "startColumn": r.min_col - 1, "endColumn": r.max_col - 1
        } for r in sheet.merged_cells.ranges]

    def _convert_cell_value(self, cell_value) -> Any:
        if isinstance(cell_value, datetime):
            return (cell_value - datetime(1899, 12, 30)).total_seconds() / 86400
        elif isinstance(cell_value, bool):
            return 1 if cell_value else 0
        elif isinstance(cell_value, pd.Timestamp):
            return (cell_value.to_pydatetime() - datetime(1899, 12, 30)).total_seconds() / 86400
        return "" if cell_value is None else cell_value

    def _extract_color(self, color) -> Optional[str]:
        if color is None: return None
        if hasattr(color, 'type'):
            if color.type == "rgb" and hasattr(color, 'rgb') and color.rgb:
                return color.rgb[-6:] if color.rgb != '00000000' else 'FFFFFF'
            elif color.type == 'theme' and hasattr(color, 'theme'):
                return ExcelUtils.theme_and_tint_to_rgb(self.theme_colors, color.theme, getattr(color, 'tint', 0))
        if hasattr(color, 'rgb') and color.rgb: return str(color.rgb)[-6:]
        return None

    def _extract_borders(self, border) -> Optional[Dict[str, Any]]:
        if border == Border(): return None
        border_data = {}
        for excel_side, univer_side in self.border_map.items():
            side = getattr(border, excel_side, None)
            if side and hasattr(side, 'style') and side.style:
                border_color = self._extract_color(getattr(side, 'color', None)) or '000000'
                border_data[univer_side] = {
                    "s": self.style_map.get(side.style, 1),
                    "cl": {"rgb": f"#{border_color}"}
                }
        return border_data if border_data else None

    def _get_row_data(self, sheet: Worksheet, max_row: int) -> Dict[str, Dict[str, Any]]:
        row_data = {}
        for r_num in range(1, max_row + 1):
            if r_num in sheet.row_dimensions:
                row_dim = sheet.row_dimensions[r_num]
                row_info = {}
                if row_dim.customHeight and row_dim.ht: row_info["h"] = row_dim.ht * 1.4
                if row_dim.hidden: row_info["hd"] = 1
                if row_info: row_data[str(r_num - 1)] = row_info
        return row_data

    def _get_column_data(self, sheet: Worksheet, max_col: int) -> Dict[str, Dict[str, Any]]:
        column_data = {}
        for c_num in range(1, max_col + 1):
            col_letter = get_column_letter(c_num)
            if col_letter in sheet.column_dimensions:
                col_dim = sheet.column_dimensions[col_letter]
                col_info = {}
                default_width = sheet.sheet_format.defaultColWidth or 8.43
                if getattr(col_dim, 'customWidth', False) and col_dim.width != default_width:
                    col_info["w"] = int(col_dim.width * 7.5)
                if col_dim.hidden: col_info["hd"] = 1
                if col_info: column_data[str(c_num - 1)] = col_info
        return column_data