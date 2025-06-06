# core/excel_handler.py
import openpyxl
import pandas as pd
from io import BytesIO
from typing import Optional, Union, Dict, Any
import logging
import tempfile
import os

logger = logging.getLogger(__name__)

class ExcelHandler:
    """Gère les opérations sur les fichiers Excel"""
    
    def __init__(self):
        self.current_workbook = None
        self.current_path = None
        self.temp_files = []
    
    def load_workbook(self, file_path: str) -> openpyxl.Workbook:
        """Charge un workbook depuis un fichier"""
        try:
            # Keep images and VBA
            wb = openpyxl.load_workbook(
                file_path, 
                data_only=False, 
                keep_vba=True,
                keep_links=False  # Disable external links
            )
            self.current_workbook = wb
            self.current_path = file_path
            logger.info(f"Workbook chargé: {file_path}")
            return wb
        except Exception as e:
            logger.error(f"Erreur chargement workbook: {str(e)}")
            raise
    
    def load_workbook_from_bytes(self, file_bytes: bytes) -> openpyxl.Workbook:
        """Charge un workbook depuis des bytes"""
        try:
            # Save bytes to temporary file to keep file reference
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                tmp_file.write(file_bytes)
                temp_path = tmp_file.name
            
            self.temp_files.append(temp_path)
            
            # Load from temp file
            wb = openpyxl.load_workbook(
                temp_path,
                data_only=False,
                keep_vba=True,
                keep_links=False
            )
            self.current_workbook = wb
            self.current_path = temp_path
            logger.info("Workbook chargé depuis bytes via fichier temporaire")
            return wb
        except Exception as e:
            logger.error(f"Erreur chargement workbook depuis bytes: {str(e)}")
            raise
    
    def save_workbook_to_bytes(self, workbook: openpyxl.Workbook) -> bytes:
        """Sauvegarde un workbook en bytes"""
        try:
            # Create a new BytesIO object
            output = BytesIO()
            
            # Save workbook
            workbook.save(output)
            
            # Get the bytes
            output.seek(0)
            result = output.getvalue()
            
            # Close BytesIO
            output.close()
            
            return result
            
        except Exception as e:
            logger.error(f"Erreur sauvegarde workbook: {str(e)}")
            # If error is due to images, try saving without them
            try:
                logger.info("Tentative de sauvegarde sans images...")
                return self._save_workbook_without_images(workbook)
            except:
                raise e
    
    def _save_workbook_without_images(self, workbook: openpyxl.Workbook) -> bytes:
        """Sauvegarde un workbook sans les images"""
        try:
            # Create a copy without images
            output = BytesIO()
            
            # Remove images from all sheets
            for sheet in workbook.worksheets:
                if hasattr(sheet, '_images'):
                    sheet._images = []
            
            # Save workbook
            workbook.save(output)
            
            # Get the bytes
            output.seek(0)
            result = output.getvalue()
            output.close()
            
            logger.warning("Workbook sauvegardé sans images")
            return result
            
        except Exception as e:
            logger.error(f"Erreur sauvegarde sans images: {str(e)}")
            raise
    
    def sheet_to_dataframe(self, workbook: openpyxl.Workbook, sheet_name: str) -> pd.DataFrame:
        """
        Converts a sheet to a pandas DataFrame.
        If the provided workbook is excel_workbook, it will have the formulas.
        If the provided workbook is displayed_excel_workbook, it will have the values (provided it has been computed)
        """
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Feuille '{sheet_name}' non trouvée")
        
        sheet = workbook[sheet_name]
        data = []

        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)
        
        if data:
            # Créer le DataFrame avec les données existantes
            df = pd.DataFrame(data)
        else:
            # Créer un DataFrame vide avec dimensions minimales
            df = pd.DataFrame()
        
        # S'assurer d'avoir au moins 20 lignes et 10 colonnes pour l'édition
        min_rows = 20
        min_cols = 10
        
        current_rows = len(df)
        current_cols = len(df.columns) if not df.empty else 0
        
        # Étendre les lignes si nécessaire
        if current_rows < min_rows:
            # Ajouter des lignes vides
            empty_rows = pd.DataFrame(index=range(current_rows, min_rows), columns=df.columns if not df.empty else range(min_cols))
            df = pd.concat([df, empty_rows], ignore_index=True)
        
        # Étendre les colonnes si nécessaire
        if current_cols < min_cols:
            for i in range(current_cols, min_cols):
                df[i] = None
        
        return df
    
    def dataframe_to_sheet(self, df: pd.DataFrame, workbook: openpyxl.Workbook, 
                        sheet_name: str, start_row: int = 1, start_col: int = 1):
        """Écrit un DataFrame dans une feuille"""
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Feuille '{sheet_name}' non trouvée")
            
        sheet = workbook[sheet_name]
        
        # D'abord, effacer toute la feuille pour éviter les données résiduelles
        # Mais préserver les formules si elles existent dans des cellules non modifiées
        for row in sheet.iter_rows():
            for cell in row:
                # Conserver les formules dans les cellules hors de la zone d'édition
                row_idx = cell.row - start_row
                col_idx = cell.column - start_col
                
                if 0 <= row_idx < len(df) and 0 <= col_idx < len(df.columns):
                    # Cette cellule sera mise à jour, on peut l'effacer
                    cell.value = None
        
        # Écrire les nouvelles données
        for r_idx, row in enumerate(df.values):
            for c_idx, value in enumerate(row):
                try:
                    cell = sheet.cell(row=start_row + r_idx, column=start_col + c_idx)
                    
                    # Gérer les différents types de valeurs
                    if pd.isna(value) or value is None or str(value).strip() == '':
                        cell.value = None
                    elif isinstance(value, str) and value.startswith('='):
                        # C'est une formule
                        cell.value = value
                    elif isinstance(value, (int, float)):
                        # Nombre
                        cell.value = value
                    else:
                        # Tout le reste en string
                        cell.value = str(value)
                        
                except Exception as e:
                    logger.warning(f"Erreur écriture cellule ({start_row + r_idx}, {start_col + c_idx}): {str(e)}")
        
        logger.info(f"DataFrame écrit dans la feuille '{sheet_name}' ({len(df)} lignes)")
        
        # Sauvegarder le workbook dans un fichier temporaire pour mettre à jour current_path
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            workbook.save(tmp.name)
            self.current_path = tmp.name
            self.temp_files.append(tmp.name)
    
    def get_sheet_info(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """Récupère les informations sur les feuilles du workbook"""
        info = {
            'sheets': [],
            'total_sheets': len(workbook.sheetnames)
        }
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_info = {
                'name': sheet_name,
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'has_formulas': False,
                'has_images': False
            }
            
            # Vérifier s'il y a des formules
            for row in sheet.iter_rows(max_row=min(100, sheet.max_row)):  # Limit scan
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        sheet_info['has_formulas'] = True
                        break
                if sheet_info['has_formulas']:
                    break
            
            # Check for images
            if hasattr(sheet, '_images') and sheet._images:
                sheet_info['has_images'] = True
            
            info['sheets'].append(sheet_info)
        
        return info
    
    def update_cell(self, workbook: openpyxl.Workbook, sheet_name: str, 
                   row: int, col: int, value: Any):
        """Met à jour une cellule spécifique"""
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Feuille '{sheet_name}' non trouvée")
        
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=col, value=value)
        logger.info(f"Cellule mise à jour: {sheet_name}!{openpyxl.utils.get_column_letter(col)}{row}")
    
    def apply_formulas_from_script(self, workbook: openpyxl.Workbook, 
                                  script_content: str) -> openpyxl.Workbook:
        """Applique les formules générées par le parseur"""
        # Cette méthode serait implémentée pour exécuter le script Python généré
        # et appliquer les résultats au workbook
        logger.info("Application des formules depuis le script")
        # TODO: Implémenter l'exécution du script
        return workbook
    
    def cleanup_temp_files(self):
        """Nettoie les fichiers temporaires"""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
                    logger.info(f"Fichier temporaire supprimé: {temp_file}")
            except Exception as e:
                logger.warning(f"Impossible de supprimer {temp_file}: {str(e)}")
        self.temp_files = []
    
    def __del__(self):
        """Destructeur pour nettoyer les fichiers temporaires"""
        self.cleanup_temp_files()