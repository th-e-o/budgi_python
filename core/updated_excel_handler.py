import openpyxl
import pandas as pd
from io import BytesIO
from typing import Optional, Any, Dict
import logging
import tempfile
import os

from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class UpdatedExcelHandler:
    """
    Manages Excel workbook operations, including the double workbook logic
    for formula computation.
    """

    def __init__(self):
        self._formula_workbook: Optional[openpyxl.Workbook] = None
        self._display_workbook: Optional[openpyxl.Workbook] = None
        self._is_stale: bool = True
        self.temp_files = []
        logger.info("ExcelHandler initialized.")

    @property
    def formula_workbook(self) -> Optional[openpyxl.Workbook]:
        """The primary workbook with formulas."""
        return self._formula_workbook

    @property
    def display_workbook(self) -> Optional[openpyxl.Workbook]:
        """The computed workbook with values. May be stale or None."""
        return self._display_workbook

    def has_workbook(self) -> bool:
        """Checks if a workbook is loaded."""
        return self._formula_workbook is not None

    def is_display_stale(self) -> bool:
        """Checks if the display workbook needs re-computation."""
        return self._is_stale

    def invalidate(self):
        """Manually marks the display workbook as stale after an external modification."""
        if not self._is_stale:
            self._is_stale = True
            logger.info("Display workbook manually invalidated.")

    def save_to_temp_file(self, workbook: openpyxl.Workbook) -> str:
        """Creates a temporary file for storing the workbook."""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()  # Close the file so it can be written to later
        workbook.save(temp_file.name)
        self.temp_files.append(temp_file.name)
        return temp_file.name

    def load_workbook_from_bytes(self, file_bytes: bytes) -> openpyxl.Workbook:
        """Loads a workbook from bytes, resetting the current state."""
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(file_bytes)
            temp_path = tmp_file.name
        self.temp_files.append(temp_path)

        wb = openpyxl.load_workbook(
            temp_path, data_only=False, keep_vba=False, keep_links=False
        )

        self._formula_workbook = wb
        self._display_workbook = None
        self._is_stale = True
        logger.info("New workbook loaded. Display workbook is stale.")
        return wb

    def replace_formula_workbook(self, new_workbook: openpyxl.Workbook):
        """
        Replaces the current formula workbook with a new one.
        Useful for tools (like BPSS) that generate a completely new workbook.
        """
        self._formula_workbook = new_workbook
        self._display_workbook = None
        self._is_stale = True
        logger.info("Formula workbook has been replaced. Display workbook is stale.")

    def apply_component_update(self, update_data: Dict[str, Any]):
        """
        Applies a cell update received from the UniverJS frontend component.
        The update format is: {"sheet": "SheetName", "range": "A1", "value": {row_idx: {col_idx: {v: value}}}, ...}
        """
        if not self.has_workbook():
            logger.warning("No workbook loaded, cannot apply update.")
            return

        sheet_name = update_data.get('sheet')
        values = update_data.get('value')

        if not sheet_name or not values or not isinstance(values, dict):
            logger.warning(f"Invalid update data received: {update_data}")
            return

        if sheet_name not in self.formula_workbook.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found in workbook. Cannot apply update.")
            return

        sheet = self.formula_workbook[sheet_name]
        logger.info(f"Applying component update to sheet: {sheet_name}")

        try:
            # The 'value' dictionary contains row and column indices (0-based)
            for row_idx_str, cols in values.items():
                for col_idx_str, cell_data in cols.items():
                    # openpyxl is 1-indexed, so we add 1
                    row = int(row_idx_str) + 1
                    col = int(col_idx_str) + 1

                    # The actual new value is nested under the 'v' key
                    new_value = cell_data.get('v')

                    # Apply the new value to the cell
                    sheet.cell(row=row, column=col).value = new_value
                    logger.debug(f"Updated cell ({row}, {col}) in '{sheet_name}' to: {new_value}")

        except (ValueError, TypeError) as e:
            logger.error(f"Error parsing update data: {e}. Data: {update_data}", exc_info=True)
        except Exception as e:
            logger.error(f"An unexpected error occurred while applying update: {e}", exc_info=True)

    def compute_display_workbook(self):
        """
        Computes formulas from the formula_workbook and updates the display_workbook.
        """
        if not self.has_workbook():
            logger.warning("Cannot compute: no formula workbook loaded.")
            return

        logger.info("Computing formulas for the display workbook...")
        try:
            self._display_workbook, error_report = self._parser.parse_and_apply(
                self.save_to_temp_file(self._formula_workbook)
            )
            print(error_report)
            self._is_stale = False
            logger.info("Display workbook computed successfully.")
        except Exception as e:
            self._is_stale = True
            logger.error(f"Failed to compute display workbook: {str(e)}", exc_info=True)
            raise

    def update_sheet_from_dataframe(self, df: pd.DataFrame, sheet_name: str):
        """
        Updates a sheet in the formula workbook based on a DataFrame from the UI.
        This invalidates the display workbook.
        """
        if not self.has_workbook():
            raise ValueError("No workbook loaded to update.")
        if sheet_name not in self.formula_workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

        sheet = self.formula_workbook[sheet_name]

        # Clear the sheet to prevent residual data
        if sheet.max_row > 0:
            sheet.delete_rows(1, sheet.max_row + 1)
        if sheet.max_column > 0:
            sheet.delete_cols(1, sheet.max_column + 1)

        df_clean = df.replace({pd.NA: None})
        for r in dataframe_to_rows(df_clean, index=False, header=True):
            sheet.append(r)

        self.invalidate()

    def sheet_to_dataframe(self, workbook: openpyxl.Workbook, sheet_name: str) -> pd.DataFrame:
        """Converts a sheet from a given workbook to a pandas DataFrame."""
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Feuille '{sheet_name}' non trouvée")

        sheet = workbook[sheet_name]
        df = pd.DataFrame(sheet.values)

        # Ensure minimum dimensions for the UI editor
        min_rows, min_cols = 20, 10
        current_rows, current_cols = df.shape

        if current_rows < min_rows:
            empty_rows = pd.DataFrame(index=range(current_rows, min_rows), columns=df.columns)
            df = pd.concat([df, empty_rows], ignore_index=True)

        if current_cols < min_cols:
            for i in range(current_cols, min_cols):
                df[i] = None

        return df

    def save_workbook_to_bytes(self, workbook: openpyxl.Workbook) -> bytes:
        """Saves a workbook to an in-memory byte stream."""
        try:
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            logger.error(f"Erreur sauvegarde workbook: {str(e)}")
            try:
                logger.warning("Tentative de sauvegarde sans images...")
                return self._save_workbook_without_images(workbook)
            except:
                raise e

    @staticmethod
    def _save_workbook_without_images(workbook: openpyxl.Workbook) -> bytes:
        """Fallback to save workbook without images if standard save fails."""
        output = BytesIO()
        for sheet in workbook.worksheets:
            if hasattr(sheet, '_images'):
                sheet._images = []
        workbook.save(output)
        output.seek(0)
        logger.warning("Workbook sauvegardé sans images")
        return output.getvalue()

    def get_sheet_info(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """Récupère les informations sur les feuilles du workbook"""
        info = {'sheets': [], 'total_sheets': len(workbook.sheetnames)}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            info['sheets'].append({
                'name': sheet_name, 'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'has_formulas': any(
                    isinstance(cell.value, str) and cell.value.startswith('=')
                    for row in sheet.iter_rows(max_row=min(100, sheet.max_row)) for cell in row
                ),
                'has_images': hasattr(sheet, '_images') and sheet._images
            })
        return info

    def cleanup_temp_files(self):
        """Cleans up temporary files created during the session."""
        for temp_file in self.temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
                    logger.info(f"Fichier temporaire supprimé: {temp_file}")
            except Exception as e:
                logger.warning(f"Impossible de supprimer {temp_file}: {str(e)}")
        self.temp_files = []

    def __del__(self):
        """Destructor to ensure cleanup."""
        self.cleanup_temp_files()