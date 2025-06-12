# app.py - Simplified version with welcome message
import streamlit as st
import asyncio
from pathlib import Path
from datetime import datetime
import logging
import openpyxl
import tempfile
import contextlib
import os
import pandas as pd

from core.updated_excel_handler import UpdatedExcelHandler
from modules.budget_mapper import BudgetMapper
from typing import List

# Configuration de la page
st.set_page_config(
    page_title="BudgiBot - Assistant Budgétaire",
    page_icon="",
    layout="wide",
    initial_sidebar_state="collapsed",  # Pas de sidebar
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': None
    }
)

# Import des modules
from core.llm_client import MistralClient
from core.file_handler import FileHandler
from core.chat_handler import ChatHandler
from modules.budget_extractor import BudgetExtractor
from modules.bpss_tool import BPSSTool
from modules.json_helper import JSONHelper
from config import config

# Import des composants UI
from ui import get_main_styles, get_javascript, MainLayout

# Configuration du logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Injection des styles CSS et JavaScript
st.markdown(get_main_styles(), unsafe_allow_html=True)
st.markdown(get_javascript(), unsafe_allow_html=True)
# CSS additionnel pour forcer la suppression des marges Streamlit
st.markdown("""
<style>
    /* Force removal of Streamlit default spacing */
    .main > div {
        padding-top: 0rem !important;
        padding-bottom: 0rem !important;
        padding-left: 1rem !important;
        padding-right: 1rem !important;
    }
    
    /* Remove header space */
    header[data-testid="stHeader"] {
        display: none !important;
    }
    
    /* Block container */
    .block-container {
        padding-top: 1rem !important;
        padding-bottom: 0rem !important;
        max-width: 100% !important;
    }
    
    /* First element fix */
    .element-container:first-child {
        margin-top: 0 !important;
    }
    
    /* Hide Streamlit menu */
    button[kind="header"] {
        display: none !important;
    }
    
    /* App container */
    .stApp {
        margin: 0 !important;
        padding: 0 !important;
    }
    
    /* Remove excessive gaps */
    div[data-testid="stVerticalBlock"] > div:not([style*="height"]) {
        gap: 0.5rem !important;
    }
    
    /* Container height fix */
    div[data-testid="stVerticalBlock"] > div[style*="height"] {
        overflow-y: auto !important;
    }
</style>
""", unsafe_allow_html=True)

# Context manager pour fichiers temporaires
@contextlib.contextmanager
def temporary_file(content, suffix=''):
    """Context manager pour fichiers temporaires"""
    fd, path = tempfile.mkstemp(suffix=suffix)
    try:
        with os.fdopen(fd, 'wb') as f:
            if isinstance(content, bytes):
                f.write(content)
            else:
                f.write(content.encode() if hasattr(content, 'encode') else bytes(content))
        yield path
    finally:
        try:
            os.unlink(path)
        except:
            pass

# Initialisation des services
@st.cache_resource
def init_services():
    """Initialise tous les services nécessaires"""
    services = {
        'llm_client': MistralClient(),
        'file_handler': FileHandler(),
        'chat_handler': ChatHandler(),
        'excel_handler': UpdatedExcelHandler(),
        'budget_extractor': BudgetExtractor(),
        'bpss_tool': BPSSTool(),
        'json_helper': JSONHelper(),
        'budget_mapper': BudgetMapper(MistralClient())
    }
    
    # Nettoyage automatique des fichiers temporaires
    import atexit
    atexit.register(lambda: services['excel_handler'].cleanup_temp_files())
    
    return services

# Message de bienvenue
WELCOME_MESSAGE = """Bonjour, 
je peux vous aider à :
• **Analyser vos fichiers Excel** - Chargez un fichier .xlsx pour commencer
• **Extraire des données budgétaires** - À partir de PDF, Word, emails ou textes
• **Utiliser l'outil BPSS** - Pour traiter vos fichiers PP-E-S, DPP18 et BUD45
"""

def init_session_state():
    """Initialise l'état de session"""
    defaults = {
        'chat_history': [],
        'current_file': None,
        'extracted_data': None,
        'is_typing': False,
        'pending_action': None,
        'json_data': None,
        'parsed_formulas': None,
        'message_input_key': 0,
        'processed_files': set(),
        'temp_files': [],
        'layout_mode': 'chat',
        'mapping_report': None,  
        'excel_tab': 'data',  # Ajouter pour gérer l'onglet actif
    }
    
    # Initialiser les valeurs par défaut
    for key, default_value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = default_value
    
    # Ajouter le message de bienvenue si première fois
    if not st.session_state.chat_history:
        st.session_state.chat_history.append({
            'role': 'assistant',
            'content': WELCOME_MESSAGE,
            'timestamp': datetime.now().strftime("%H:%M"),
            'type': 'welcome'
        })

def cleanup_temp_files():
    """Nettoie les fichiers temporaires"""
    for temp_file in st.session_state.get('temp_files', []):
        try:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
                logger.info(f"Fichier temporaire supprimé: {temp_file}")
        except Exception as e:
            logger.warning(f"Impossible de supprimer {temp_file}: {str(e)}")
    st.session_state.temp_files = []

# Gestionnaires d'événements
async def handle_message_send(message: str):
    """Gère l'envoi d'un message"""
    # Ajouter le message utilisateur
    st.session_state.chat_history.append({
        'role': 'user',
        'content': message,
        'timestamp': datetime.now().strftime("%H:%M")
    })
    
    # Activer l'indicateur de frappe
    st.session_state.is_typing = True
    st.rerun()

async def process_message():
    """Traite le message avec le LLM"""
    try:
        # Préparer les messages pour l'API
        api_messages = services['chat_handler'].filter_messages_for_api(
            st.session_state.chat_history
        )
        
        # Obtenir la réponse
        response = await services['llm_client'].chat(api_messages)
        
        if response:
            # Ajouter la réponse
            st.session_state.chat_history.append({
                'role': 'assistant',
                'content': response,
                'timestamp': datetime.now().strftime("%H:%M")
            })
            
            # Détecter si BPSS est mentionné
            last_message = st.session_state.chat_history[-2]['content'].lower()
            if any(keyword in last_message for keyword in ['bpss', 'mesures catégorielles']):
                st.session_state.excel_tab = 'tools'
                st.session_state.layout_mode = 'split'
                
    except Exception as e:
        logger.error(f"Erreur traitement message: {str(e)}")
        st.session_state.chat_history.append({
            'role': 'assistant',
            'content': "❌ Désolé, une erreur s'est produite. Pouvez-vous reformuler ?",
            'timestamp': datetime.now().strftime("%H:%M"),
            'error': True
        })
    
    st.session_state.is_typing = False
    st.rerun()

async def handle_file_upload(uploaded_file):
    """Gère l'upload d'un fichier"""
    # Éviter les doublons
    file_key = f"{uploaded_file.name}_{uploaded_file.size}"
    if file_key in st.session_state.processed_files:
        return
    
    st.session_state.processed_files.add(file_key)
    
    # Ajouter le message d'upload
    st.session_state.chat_history.append({
        'role': 'user',
        'content': f"📎 Fichier envoyé : {uploaded_file.name}",
        'timestamp': datetime.now().strftime("%H:%M"),
        'file_name': uploaded_file.name,
        'file_size': uploaded_file.size,
        'file_key': file_key
    })
    
    st.session_state.is_typing = True
    st.rerun()

async def process_file(uploaded_file):
    """Traite le fichier uploadé"""
    try:
        file_content = uploaded_file.getbuffer()
        suffix = Path(uploaded_file.name).suffix
        
        with temporary_file(file_content, suffix=suffix) as temp_path:
            # Lire le contenu
            content = services['file_handler'].read_file(temp_path, uploaded_file.name)
            
            # Stocker les informations
            st.session_state.current_file = {
                'name': uploaded_file.name,
                'content': content,
                'type': suffix[1:].lower(),
                'raw_bytes': file_content,
                'size': uploaded_file.size
            }
            
            # Stocker le contenu (caché)
            st.session_state.chat_history.append({
                'role': 'system',
                'content': content,
                'meta': 'file_content',
                'file_name': uploaded_file.name,
                'timestamp': datetime.now().strftime("%H:%M")
            })
            
            # Traitement spécifique
            if uploaded_file.name.endswith('.xlsx'):
                wb = services['excel_handler'].load_workbook_from_bytes(file_content)
                st.session_state.layout_mode = 'split'

                response = f"✅ J'ai chargé votre fichier Excel '{uploaded_file.name}'. Il contient {len(st.session_state.excel_workbook.sheetnames)} feuilles. Vous pouvez maintenant :\n\n• Visualiser et éditer les données dans l'onglet Excel\n• Extraire les données budgétaires\n• Utiliser l'outil BPSS pour les mesures catégorielles"
                
            elif uploaded_file.name.endswith('.json'):
                import json
                st.session_state.json_data = json.loads(content)
                response = f"✅ Fichier JSON de configuration chargé. Il contient {len(st.session_state.json_data.get('tags', []))} tags pour le mapping automatique."
                
            else:
                # Résumé pour autres types
                preview = content[:500] + "..." if len(content) > 500 else content
                response = f"✅ J'ai bien reçu votre fichier '{uploaded_file.name}'. Voici un aperçu :\n\n{preview}\n\nQue souhaitez-vous faire avec ce fichier ?"
            
            # Ajouter la réponse
            st.session_state.chat_history.append({
                'role': 'assistant',
                'content': response,
                'timestamp': datetime.now().strftime("%H:%M")
            })
        
    except Exception as e:
        logger.error(f"Erreur traitement fichier: {str(e)}")
        st.session_state.chat_history.append({
            'role': 'assistant',
            'content': f"❌ Erreur lors du traitement : {str(e)}",
            'timestamp': datetime.now().strftime("%H:%M"),
            'error': True
        })
    
    st.session_state.is_typing = False
    st.rerun()

def handle_tool_action(action: dict):
    """Gère les actions des outils"""
    action_type = action.get('action')
    
    if action_type == 'clear_history':
        # Réinitialiser la conversation
        st.session_state.chat_history = [{
            'role': 'assistant',
            'content': WELCOME_MESSAGE,
            'timestamp': datetime.now().strftime("%H:%M"),
            'type': 'welcome'
        }]
        st.session_state.processed_files = set()
        st.session_state.current_file = None
        st.session_state.excel_workbook = None
        st.session_state.extracted_data = None
        cleanup_temp_files()
        st.success("✨ Conversation réinitialisée")
        
    elif action_type == 'extract_budget':
        asyncio.run(extract_budget_data())
        
    elif action_type == 'process_bpss':
        asyncio.run(process_bpss(action.get('data')))
        
    elif action_type == 'map_budget_cells':
        asyncio.run(map_budget_to_cells())

async def extract_budget_data():
    """Extrait les données budgétaires"""
    if not st.session_state.current_file:
        st.error("❌ Aucun fichier chargé")
        return
    
    with st.spinner("Extraction en cours..."):
        try:
            content = st.session_state.current_file['content']
            file_name = st.session_state.current_file['name']
            
            # Limiter la taille
            if len(content) > 10000:
                content = content[:10000] + "\n\n[... contenu tronqué ...]"
            
            # Extraire
            data = await services['budget_extractor'].extract(
                content, 
                services['llm_client']
            )
            
            if data:
                st.session_state.extracted_data = data
                st.session_state.excel_tab = 'analysis'
                
                # Message de succès
                st.session_state.chat_history.append({
                    'role': 'assistant',
                    'content': f"✅ J'ai extrait **{len(data)} entrées budgétaires** du fichier '{file_name}'.\n\nRendez-vous dans l'onglet 'Extraction' pour visualiser et éditer les données.",
                    'timestamp': datetime.now().strftime("%H:%M")
                })
                st.success(f"✅ {len(data)} entrées extraites!")
            else:
                st.warning("⚠️ Aucune donnée budgétaire trouvée")
                
        except Exception as e:
            logger.error(f"Erreur extraction: {str(e)}")
            st.error(f"❌ Erreur: {str(e)}")

async def process_bpss(data: dict):
    """Traite les fichiers BPSS"""
    try:
        progress = st.progress(0, text="Traitement BPSS...")
        
        temp_files = []
        temp_paths = {}
        
        try:
            # Sauvegarder les fichiers temporairement
            for key, file in data['files'].items():
                # Créer un fichier temporaire
                fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
                temp_files.append(temp_path)  # Garder la trace pour nettoyage
                
                # Écrire le contenu
                with os.fdopen(fd, 'wb') as tmp:
                    tmp.write(file.getbuffer())
                
                temp_paths[key] = temp_path
                logger.info(f"Fichier temporaire créé: {key} -> {temp_path}")
            
            progress.progress(50, text="Application des données...")

            target_wb = services['excel_handler'].formula_workbook or openpyxl.Workbook()

            # Traiter avec les fichiers temporaires
            result_wb = services['bpss_tool'].process_files(
                ppes_path=temp_paths['ppes'],
                dpp18_path=temp_paths['dpp18'],
                bud45_path=temp_paths['bud45'],
                year=data['year'],
                ministry_code=data['ministry'],
                program_code=data['program'],
                target_workbook=st.session_state.excel_workbook or openpyxl.Workbook()
            )

            services['excel_handler'].replace_formula_workbook(result_wb)
            progress.progress(100, text="Terminé!")
            
            # Message de succès
            st.session_state.chat_history.append({
                'role': 'assistant',
                'content': f"✅ Traitement BPSS terminé!\n\nJ'ai intégré les données pour:\n• Année: {data['year']}\n• Ministère: {data['ministry']}\n• Programme: {data['program']}\n\nLes feuilles ont été ajoutées à votre fichier Excel.",
                'timestamp': datetime.now().strftime("%H:%M")
            })
            
            st.success("✅ Traitement BPSS réussi!")

            # IMPORTANT : Forcer le rechargement
            st.session_state.excel_workbook = result_wb

            # Sauvegarder temporairement pour l'affichage des valeurs
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                result_wb.save(tmp.name)
                services['excel_handler'].current_path = tmp.name
                temp_files.append(tmp.name)

            # Message de succès avec détails
            st.success(f"""
            ✅ Traitement BPSS terminé!
            - Année: {data['year']}
            - Ministère: {data['ministry']}
            - Programme: {data['program']}

            Feuilles ajoutées: Données PP-E-S, INF DPP 18, INF BUD 45
            """)

            # Basculer vers l'onglet Excel
            st.session_state.layout_mode = 'excel'

            # FORCER LE RAFRAÎCHISSEMENT
            st.rerun()
            
        finally:
            # Nettoyer les fichiers temporaires
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.unlink(temp_file)
                        logger.info(f"Fichier temporaire supprimé: {temp_file}")
                except Exception as e:
                    logger.warning(f"Impossible de supprimer {temp_file}: {str(e)}")
        
    except Exception as e:
        logger.error(f"Erreur BPSS: {str(e)}")
        st.error(f"❌ Erreur: {str(e)}")


async def map_budget_to_cells():
    """Mappe les données aux cellules Excel avec gestion du rapport"""
    # Ensure that extracted data and JSON data are available
    if not st.session_state.get('extracted_data') or not st.session_state.get('json_data'):
        st.error("❌ Données manquantes pour le mapping")
        return

    # Ensure a workbook is loaded in the handler
    if not services['excel_handler'].has_workbook():
        st.error("❌ Aucun fichier Excel n'est chargé pour le mapping.")
        return

    with st.spinner("Mapping en cours..."):
        try:
            mapper = services['budget_mapper']
            tags = services['json_helper'].get_tags_for_mapping(st.session_state.json_data)
            
            # Créer une barre de progression
            progress_bar = st.progress(0)
            progress_text = st.empty()
            
            def update_progress(percent, text):
                progress_bar.progress(int(percent))
                progress_text.text(text)
            
            # Mapper avec callback de progression
            mapping = await mapper.map_entries_to_cells(
                st.session_state.extracted_data,
                tags,
                progress_callback=update_progress
            )
            
            # Nettoyer la barre de progression
            progress_bar.empty()
            progress_text.empty()
            
            if mapping:
                # Enrichir les entrées avec le mapping
                entries_df = pd.DataFrame(st.session_state.extracted_data)
                enriched_df = mapper.enrich_entries_with_mapping(entries_df, mapping)
                
                # Mettre à jour les données extraites enrichies
                st.session_state.extracted_data = enriched_df.to_dict('records')
                
                # Générer le rapport de mapping
                report = mapper.generate_mapping_report(mapping, entries_df)
                st.session_state.mapping_report = report
                
                # Appliquer au workbook
                success, errors = mapper.apply_mapping_to_excel(
                    services['excel_handler'].formula_workbook,
                    mapping,
                    entries_df
                )

                # After modification, invalidate the display workbook
                services['excel_handler'].invalidate()
                
                # Afficher les résultats SANS colonnes imbriquées
                if success > 0:
                    # Message de succès simple
                    st.success(f"""
                    ✅ Mapping terminé avec succès!
                    
                    **Résultats:**
                    - {success} cellules mises à jour
                    - {report['summary']['mapping_rate']:.1f}% de taux de mapping
                    - {report['summary']['average_confidence']:.1%} de confiance moyenne
                    
                    👉 Consultez l'interface de vérification dans l'onglet Excel pour valider les mappings.
                    """)
                    
                    # Message dans le chat
                    st.session_state.chat_history.append({
                        'role': 'assistant',
                        'content': f"✅ Mapping terminé!\n\n• **{success}** cellules mises à jour\n• **{report['summary']['mapping_rate']:.1f}%** de taux de mapping\n• **{report['summary']['average_confidence']:.1%}** de confiance moyenne\n\nConsultez l'interface de vérification dans l'onglet Excel pour valider les mappings.",
                        'timestamp': datetime.now().strftime("%H:%M")
                    })
                    
                    # Basculer vers la vue Excel pour voir les résultats
                    st.session_state.layout_mode = 'excel'
                    st.rerun()
                    
                if errors:
                    # Afficher les erreurs dans un expander
                    with st.expander("⚠️ Problèmes rencontrés", expanded=False):
                        for error in errors[:10]:  # Limiter à 10 erreurs
                            st.warning(error)
                        if len(errors) > 10:
                            st.warning(f"... et {len(errors) - 10} autres problèmes")
            else:
                st.warning("⚠️ Aucun mapping n'a pu être établi")
                            
        except Exception as e:
            logger.error(f"Erreur mapping: {str(e)}")
            st.error(f"❌ Erreur lors du mapping: {str(e)}")

# Initialisation des services
services = init_services()

def main():
    """Fonction principale"""
    # Initialiser l'état
    init_session_state()
    
    # Nettoyer au démarrage
    if 'startup_cleanup' not in st.session_state:
        cleanup_temp_files()
        st.session_state.startup_cleanup = True
    
    # Traiter les messages/fichiers en attente
    if st.session_state.is_typing and st.session_state.chat_history:
        last_msg = st.session_state.chat_history[-1]
        
        if last_msg['role'] == 'user':
            if last_msg['content'].startswith("📎"):
                # Fichier à traiter
                file_key = last_msg.get('file_key')
                if file_key:
                    # Chercher le fichier dans les widgets
                    for widget_key in st.session_state:
                        if widget_key.startswith('file_upload_'):
                            file = st.session_state.get(widget_key)
                            if file and f"{file.name}_{file.size}" == file_key:
                                asyncio.run(process_file(file))
                                break
            else:
                # Message texte
                asyncio.run(process_message())
    
    # Traiter les actions en attente
    if st.session_state.get('pending_action'):
        action = st.session_state.pending_action
        st.session_state.pending_action = None
        
        if action['type'] == 'extract_budget':
            asyncio.run(extract_budget_data())
    
    # Créer et rendre l'interface
    layout = MainLayout(services)
    layout.render(
        on_message_send=lambda msg: asyncio.run(handle_message_send(msg)),
        on_file_upload=lambda file: asyncio.run(handle_file_upload(file)),
        on_tool_action=handle_tool_action
    )

if __name__ == "__main__":
    main()