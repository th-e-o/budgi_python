import openpyxl
import pandas as pd
import numpy as np
import re
from typing import Dict, List, Tuple, Optional, Any, Set
import concurrent.futures
from dataclasses import dataclass
import logging
from functools import lru_cache
import gc
from tqdm import tqdm
from collections import defaultdict
import ast
import operator
import datetime

logger = logging.getLogger(__name__)

@dataclass
class ParserConfig:
    chunk_size: int = 800
    max_memory_mb: int = 1024
    workers: int = 4
    cache_enabled: bool = True
    progress_enabled: bool = True

@dataclass
class FormulaCell:
    sheet: str
    address: str
    row: int
    col: int
    formula: str
    python_code: Optional[str] = None
    dependencies: List[str] = None
    value: Any = None
    error: Optional[str] = None

class ExcelFormulaParser:
    """Parseur de formules Excel optimisé - implémentation complète basée sur le parseur R"""
    
    def __init__(self, config: Optional[ParserConfig] = None):
        self.config = config or ParserConfig()
        self._ref_cache = {}
        self._formula_converters = self._init_converters()
        self._named_ranges = {}
        self._sheets_data = {}
        
    def _init_converters(self) -> Dict:
        """Initialise les convertisseurs de formules"""
        return {
            'SUM': self._convert_sum,
            'MAX': self._convert_max,
            'MIN': self._convert_min,
            'AVERAGE': self._convert_average,
            'COUNT': self._convert_count,
            'COUNTA': self._convert_counta,
            'ROUND': self._convert_round,
            'INT': self._convert_int,
            'IF': self._convert_if,
            'SUMIF': self._convert_sumif,
            'IFERROR': self._convert_iferror,
            'TEXT': self._convert_text,
            'CONCATENATE': self._convert_concatenate,
            'AND': self._convert_and,
            'OR': self._convert_or,
            'NOT': self._convert_not,
            'ISBLANK': self._convert_isblank,
            'DATE': self._convert_date,
            'OFFSET': self._convert_offset,
            'INDIRECT': self._convert_indirect,
            'VLOOKUP': self._convert_vlookup,
            'INDEX': self._convert_index,
            'MATCH': self._convert_match,
            'LEN': self._convert_len,
            'TRIM': self._convert_trim,
            'UPPER': self._convert_upper,
            'LOWER': self._convert_lower,
            'MID': self._convert_mid,
            'SUBSTITUTE': self._convert_substitute,
            'ABS': self._convert_abs,
            'SQRT': self._convert_sqrt,
            'POWER': self._convert_power,
            'MOD': self._convert_mod,
        }
    
    @lru_cache(maxsize=10000)
    def excel_col_to_num(self, col_str: str) -> int:
        """Convertit une colonne Excel (A, AB, etc.) en numéro (1-based)"""
        if not col_str:
            return 1
        
        result = 0
        for char in col_str:
            result = result * 26 + ord(char) - ord('A') + 1
        return result
    
    def parse_excel_file(self, file_path: str, emit_script: bool = True) -> Dict:
        """Parse un fichier Excel et extrait les formules"""
        logger.info(f"Parsing Excel file: {file_path}")
        
        # Charger le workbook
        wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=False)
        
        # Charger les named ranges
        self._load_named_ranges(wb)
        
        # Charger toutes les données des feuilles pour référence
        self._load_sheets_data(wb)
        
        # Extraire toutes les formules
        all_formulas = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_formulas = self._extract_sheet_formulas(sheet, sheet_name)
            all_formulas.extend(sheet_formulas)
        
        wb.close()
        
        logger.info(f"Found {len(all_formulas)} formulas")
        
        # Filtrer les erreurs Excel
        filtered_formulas = self._filter_excel_errors(all_formulas)
        logger.info(f"Filtered to {len(filtered_formulas)} valid formulas")
        
        # Extraire les dépendances
        for formula in filtered_formulas:
            formula.dependencies = self._extract_dependencies(formula)
        
        # Ordonner les formules selon les dépendances
        ordered_formulas = self._topological_sort(filtered_formulas)
        
        # Convertir les formules
        if self.config.progress_enabled:
            converted_formulas = self._convert_formulas_with_progress(ordered_formulas)
        else:
            converted_formulas = self._convert_formulas_batch(ordered_formulas)
        
        # Générer le script si demandé
        script_file = None
        if emit_script and converted_formulas:
            script_file = self._generate_python_script(converted_formulas, file_path)
        
        # Statistiques
        success_count = sum(1 for f in converted_formulas if f.python_code and not f.error)
        error_count = len(converted_formulas) - success_count
        
        return {
            'formulas': converted_formulas,
            'script_file': script_file,
            'statistics': {
                'total': len(all_formulas),
                'success': success_count,
                'errors': error_count,
                'success_rate': round(100 * success_count / len(all_formulas), 1) if all_formulas else 0
            }
        }
    
    def _load_named_ranges(self, wb):
        """Charge les plages nommées du workbook"""
        self._named_ranges = {}
        if hasattr(wb, 'defined_names'):
            for name in wb.defined_names:
                # Extraire la référence
                if hasattr(name, 'value'):
                    self._named_ranges[name.name] = name.value
    
    def _load_sheets_data(self, wb):
        """Charge toutes les données des feuilles pour référence"""
        self._sheets_data = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Créer un DataFrame pour chaque feuille
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            if data:
                self._sheets_data[sheet_name] = pd.DataFrame(data)
            else:
                self._sheets_data[sheet_name] = pd.DataFrame()
    
    def _extract_sheet_formulas(self, sheet, sheet_name: str) -> List[FormulaCell]:
        """Extrait les formules d'une feuille"""
        formulas = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = FormulaCell(
                        sheet=sheet_name,
                        address=cell.coordinate,
                        row=cell.row,
                        col=cell.column,
                        formula=cell.value[1:]  # Enlever le '='
                    )
                    formulas.append(formula)
        
        return formulas
    
    def _filter_excel_errors(self, formulas: List[FormulaCell]) -> List[FormulaCell]:
        """Filtre les formules contenant des erreurs Excel"""
        error_patterns = [
            '#REF!', '#N/A', '#VALUE!', '#DIV/0!', 
            '#NAME?', '#NULL!', '#NUM!'
        ]
        
        filtered = []
        for formula in formulas:
            has_error = any(error in formula.formula for error in error_patterns)
            if not has_error:
                filtered.append(formula)
            else:
                formula.error = "Contains Excel error"
                filtered.append(formula)  # On garde mais on marque l'erreur
        
        return filtered
    
    def _extract_dependencies(self, formula: FormulaCell) -> List[str]:
        """Extrait les dépendances d'une formule"""
        dependencies = []
        
        # Pattern pour les références de cellules
        # Sheet!Cell ou Cell
        pattern = r"(?:(?:'[^']+'|[A-Za-z0-9_]+)!)?([A-Z]+[0-9]+)(?::([A-Z]+[0-9]+))?"
        
        matches = re.finditer(pattern, formula.formula)
        for match in matches:
            full_match = match.group(0)
            
            # Déterminer la feuille
            if '!' in full_match:
                sheet_part = full_match.split('!')[0].strip("'")
            else:
                sheet_part = formula.sheet
            
            # Extraire les cellules
            cell1 = match.group(1)
            cell2 = match.group(2)
            
            if cell2:  # C'est une plage
                # Calculer toutes les cellules dans la plage
                col1 = re.match(r'([A-Z]+)', cell1).group(1)
                row1 = int(re.match(r'[A-Z]+(\d+)', cell1).group(1))
                col2 = re.match(r'([A-Z]+)', cell2).group(1)
                row2 = int(re.match(r'[A-Z]+(\d+)', cell2).group(1))
                
                col1_num = self.excel_col_to_num(col1)
                col2_num = self.excel_col_to_num(col2)
                
                for r in range(min(row1, row2), max(row1, row2) + 1):
                    for c in range(min(col1_num, col2_num), max(col1_num, col2_num) + 1):
                        col_letter = self._num_to_col(c)
                        dependencies.append(f"{sheet_part}!{col_letter}{r}")
            else:
                dependencies.append(f"{sheet_part}!{cell1}")
        
        # Dédupliquer
        return list(set(dependencies))
    
    def _num_to_col(self, num: int) -> str:
        """Convertit un numéro de colonne en lettre Excel"""
        result = ""
        while num > 0:
            num -= 1
            result = chr(num % 26 + ord('A')) + result
            num //= 26
        return result
    
    def _topological_sort(self, formulas: List[FormulaCell]) -> List[FormulaCell]:
        """Trie les formules selon leurs dépendances"""
        # Créer un mapping des adresses vers les formules
        formula_map = {}
        for formula in formulas:
            key = f"{formula.sheet}!{formula.address}"
            formula_map[key] = formula
        
        # Créer le graphe de dépendances
        graph = defaultdict(list)
        in_degree = defaultdict(int)
        
        for formula in formulas:
            key = f"{formula.sheet}!{formula.address}"
            in_degree[key] = 0
        
        for formula in formulas:
            key = f"{formula.sheet}!{formula.address}"
            for dep in formula.dependencies:
                if dep in formula_map:
                    graph[dep].append(key)
                    in_degree[key] += 1
        
        # Tri topologique
        queue = [key for key in in_degree if in_degree[key] == 0]
        sorted_formulas = []
        
        while queue:
            current = queue.pop(0)
            if current in formula_map:
                sorted_formulas.append(formula_map[current])
            
            for neighbor in graph[current]:
                in_degree[neighbor] -= 1
                if in_degree[neighbor] == 0:
                    queue.append(neighbor)
        
        # Ajouter les formules restantes (cycles potentiels)
        for formula in formulas:
            if formula not in sorted_formulas:
                sorted_formulas.append(formula)
        
        return sorted_formulas
    
    def _convert_formulas_with_progress(self, formulas: List[FormulaCell]) -> List[FormulaCell]:
        """Convertit les formules avec une barre de progression"""
        with tqdm(total=len(formulas), desc="Converting formulas") as pbar:
            for formula in formulas:
                self._convert_single_formula(formula)
                pbar.update(1)
        
        return formulas
    
    def _convert_formulas_batch(self, formulas: List[FormulaCell]) -> List[FormulaCell]:
        """Convertit un batch de formules"""
        for formula in formulas:
            self._convert_single_formula(formula)
        
        return formulas
    
    def _convert_single_formula(self, formula_cell: FormulaCell):
        """Convertit une formule Excel en code Python"""
        try:
            formula_cell.python_code = self._parse_formula(
                formula_cell.formula, 
                formula_cell.sheet
            )
        except Exception as e:
            formula_cell.error = str(e)
            formula_cell.python_code = f"# Error: {str(e)}"
    
    def _parse_formula(self, formula: str, current_sheet: str) -> str:
        """Parse et convertit une formule Excel en Python - version refaite"""
        # Nettoyer la formule
        formula = formula.strip()
        
        # Log pour debug
        logger.debug(f"Parsing formula: '{formula}'")
        
        # Cas de base : constantes
        if not formula:
            return '""'
        
        # Nombres
        if self._is_number(formula):
            return formula.replace(',', '.')  # Gérer les virgules décimales
        
        # Chaînes
        if formula.startswith('"') and formula.endswith('"'):
            return formula
        
        # Booléens
        if formula.upper() in ['TRUE', 'FALSE']:
            return formula.title()
        
        # Pourcentages
        if formula.endswith('%') and self._is_number(formula[:-1]):
            return f"({formula[:-1]}/100)"
        
        # Références de cellules (avant de chercher les fonctions!)
        if self._is_cell_reference(formula):
            return self._convert_cell_reference(formula, current_sheet)
        
        # Vérifier si c'est une fonction (doit avoir des parenthèses correspondantes)
        # Pattern : NOM_FONCTION(...)
        func_match = re.match(r'^([A-Z_][A-Z0-9_]*)\s*\(', formula, re.IGNORECASE)
        if func_match:
            # C'est potentiellement une fonction
            func_name = func_match.group(1)
            
            # Trouver la parenthèse fermante correspondante
            paren_start = func_match.end() - 1  # Position de la parenthèse ouvrante
            paren_depth = 1
            paren_end = paren_start
            
            for i in range(paren_start + 1, len(formula)):
                if formula[i] == '(':
                    paren_depth += 1
                elif formula[i] == ')':
                    paren_depth -= 1
                    if paren_depth == 0:
                        paren_end = i
                        break
            
            if paren_depth == 0 and paren_end == len(formula) - 1:
                # C'est une fonction complète
                return self._parse_function(formula, current_sheet)
        
        # Si ce n'est pas une fonction, chercher les opérateurs
        # Important : traiter les parenthèses d'abord
        if formula.startswith('(') and formula.endswith(')'):
            # Vérifier que les parenthèses sont bien appariées
            depth = 0
            for i, char in enumerate(formula):
                if char == '(':
                    depth += 1
                elif char == ')':
                    depth -= 1
                    if depth == 0 and i < len(formula) - 1:
                        # Les parenthèses ne sont pas englobantes
                        break
            
            if depth == 0 and i == len(formula) - 1:
                # Les parenthèses englobent toute l'expression
                inner = formula[1:-1]
                inner_parsed = self._parse_formula(inner, current_sheet)
                return f"({inner_parsed})"
        
        # Chercher les opérateurs binaires (ordre de priorité)
        # Plus basse priorité vers plus haute priorité
        operator_groups = [
            ['&'],                    # Concaténation
            ['=', '<>', '<=', '>=', '<', '>'],  # Comparaisons
            ['+', '-'],              # Addition, soustraction
            ['*', '/'],              # Multiplication, division
            ['^']                    # Puissance
        ]
        
        for ops in operator_groups:
            result = self._try_split_binary(formula, ops, current_sheet)
            if result:
                return result
        
        # Si c'est un nom défini
        if formula in self._named_ranges:
            return self._convert_cell_reference(self._named_ranges[formula], current_sheet)
        
        # Si on arrive ici, c'est quelque chose qu'on ne reconnaît pas
        logger.warning(f"Unrecognized formula element: '{formula}'")
        return f'"{formula}"'
    
    def _is_number(self, s: str) -> bool:
        """Vérifie si une chaîne est un nombre"""
        try:
            # Gérer les nombres avec virgule comme séparateur décimal
            s_normalized = s.replace(',', '.')
            float(s_normalized)
            return True
        except ValueError:
            return False
        
    def _is_cell_reference(self, s: str) -> bool:
        """Vérifie si c'est une référence de cellule"""
        # Enlever les $ pour la vérification
        s_clean = s.replace('$', '')
        
        # Simple référence A1 ou plage A1:B2
        if re.match(r'^[A-Z]+\d+(?::[A-Z]+\d+)?$', s_clean, re.IGNORECASE):
            return True
        # Avec feuille Sheet!A1 ou 'Sheet'!A1
        if re.match(r"^(?:'[^']+'|[A-Za-z0-9_]+)![A-Z]+\d+(?::[A-Z]+\d+)?$", s_clean, re.IGNORECASE):
            return True
        return False
    
    def _try_split_binary(self, formula: str, operators: List[str], current_sheet: str) -> Optional[str]:
        """Divise sur un opérateur binaire en respectant les parenthèses"""
        depth = 0
        in_string = False
        quote_char = None
        
        # Parcourir de droite à gauche pour respecter l'associativité
        for i in range(len(formula) - 1, -1, -1):
            char = formula[i]
            
            # Gestion des chaînes (à l'envers)
            if not in_string:
                if char in ['"', "'"]:
                    # Vérifier que ce n'est pas échappé
                    if i > 0 and formula[i-1] != '\\':
                        in_string = True
                        quote_char = char
            else:
                if char == quote_char and (i == 0 or formula[i-1] != '\\'):
                    in_string = False
                    quote_char = None
                continue
            
            # Gestion des parenthèses
            if char == ')':
                depth += 1
            elif char == '(':
                depth -= 1
            
            # Si on est au niveau racine
            if depth == 0 and not in_string:
                for op in operators:
                    # Vérifier si l'opérateur est présent à cette position
                    if i >= len(op) - 1 and formula[i - len(op) + 1:i + 1] == op:
                        # Vérifier que ce n'est pas partie d'un autre opérateur
                        if op in ['<', '>'] and i < len(formula) - 1 and formula[i + 1] == '=':
                            continue
                        
                        left = formula[:i - len(op) + 1].strip()
                        right = formula[i + 1:].strip()
                        
                        if left and right:
                            left_py = self._parse_formula(left, current_sheet)
                            right_py = self._parse_formula(right, current_sheet)
                            
                            # Convertir l'opérateur
                            if op == '&':
                                return f'str({left_py}) + str({right_py})'
                            elif op == '<>':
                                return f'({left_py} != {right_py})'
                            elif op == '=':
                                return f'({left_py} == {right_py})'
                            elif op == '^':
                                return f'({left_py} ** {right_py})'
                            else:
                                return f'({left_py} {op} {right_py})'
        
        return None
    
    def _parse_function(self, formula: str, current_sheet: str) -> str:
        """Parse une fonction Excel - version améliorée pour expressions complexes"""
        # Gérer les espaces
        formula = formula.strip()
        
        # Trouver le nom de la fonction et les parenthèses
        paren_pos = formula.find('(')
        if paren_pos == -1:
            return f'# No parentheses in function: {formula}'
        
        func_name = formula[:paren_pos].strip().upper()
        
        # Vérifier que les parenthèses sont équilibrées
        if not formula.endswith(')'):
            return f'# Unbalanced parentheses: {formula}'
        
        # Extraire le contenu entre parenthèses
        args_str = formula[paren_pos + 1:-1]
        
        # Parser les arguments
        try:
            args = self._split_arguments(args_str)
        except Exception as e:
            logger.error(f"Error splitting arguments for {func_name}: {str(e)}")
            return f'# Error parsing arguments: {formula}'
        
        # Debug pour voir ce qui se passe
        logger.debug(f"Function {func_name} with {len(args)} arguments")
        for i, arg in enumerate(args):
            logger.debug(f"  Arg {i}: '{arg}'")
        
        # Convertir chaque argument
        converted_args = []
        for i, arg in enumerate(args):
            try:
                # Pour debug
                logger.debug(f"Converting arg {i} of {func_name}: '{arg}'")
                converted = self._parse_formula(arg.strip(), current_sheet)
                logger.debug(f"  -> '{converted}'")
                converted_args.append(converted)
            except Exception as e:
                logger.error(f"Error converting argument {i} of {func_name}: {arg}")
                logger.error(f"  Error: {str(e)}")
                converted_args.append(f"# Error: {arg}")
        
        # Utiliser le convertisseur approprié
        if func_name in self._formula_converters:
            try:
                result = self._formula_converters[func_name](converted_args, current_sheet, args)
                logger.debug(f"Converted {func_name} to: {result}")
                return result
            except Exception as e:
                logger.error(f"Error in converter for {func_name}: {str(e)}")
                return f"# Error converting {func_name}: {str(e)}"
        else:
            # Fonction non supportée
            logger.warning(f"Unsupported function: {func_name}")
            # Pour les fonctions non supportées, retourner un appel générique
            return f"{func_name.lower()}({', '.join(converted_args)})"
    
    def _split_arguments(self, args_str: str) -> List[str]:
        """Divise les arguments d'une fonction - version robuste"""
        if not args_str.strip():
            return []
        
        args = []
        current_arg = ""
        paren_depth = 0
        bracket_depth = 0
        in_string = False
        quote_char = None
        i = 0
        
        while i < len(args_str):
            char = args_str[i]
            
            # Gestion des chaînes
            if not in_string and char in ['"', "'"]:
                in_string = True
                quote_char = char
                current_arg += char
            elif in_string and char == quote_char:
                # Vérifier si c'est échappé
                if i + 1 < len(args_str) and args_str[i + 1] == quote_char:
                    current_arg += char + args_str[i + 1]
                    i += 1
                else:
                    in_string = False
                    quote_char = None
                    current_arg += char
            elif not in_string:
                # Gestion des parenthèses et crochets
                if char == '(':
                    paren_depth += 1
                elif char == ')':
                    paren_depth -= 1
                elif char == '[':
                    bracket_depth += 1
                elif char == ']':
                    bracket_depth -= 1
                elif char == ',' and paren_depth == 0 and bracket_depth == 0:
                    # C'est un séparateur d'arguments
                    args.append(current_arg.strip())
                    current_arg = ""
                    i += 1
                    continue
                
                current_arg += char
            else:
                current_arg += char
            
            i += 1
        
        # Ajouter le dernier argument
        if current_arg.strip():
            args.append(current_arg.strip())
        
        return args
    
        
        if current_arg.strip():
            args.append(current_arg.strip())
        
        return args
    
    def _safe_cell_access(self, df: pd.DataFrame, row: int, col: int):
        """Accès sécurisé à une cellule du DataFrame"""
        try:
            if row < 0 or col < 0:
                return None
            if row >= len(df) or col >= len(df.columns):
                return None
            return df.iloc[row, col]
        except:
            return None

    def _safe_range_access(self, df: pd.DataFrame, row1: int, col1: int, row2: int, col2: int):
        """Accès sécurisé à une plage du DataFrame"""
        try:
            # S'assurer que les indices sont dans les limites
            row1 = max(0, min(row1, len(df) - 1))
            row2 = max(0, min(row2, len(df) - 1))
            col1 = max(0, min(col1, len(df.columns) - 1))
            col2 = max(0, min(col2, len(df.columns) - 1))
            
            # S'assurer que row1 <= row2 et col1 <= col2
            if row1 > row2:
                row1, row2 = row2, row1
            if col1 > col2:
                col1, col2 = col2, col1
                
            return df.iloc[row1:row2+1, col1:col2+1]
        except:
            return pd.DataFrame()

    def _convert_cell_reference(self, ref: str, current_sheet: str) -> str:
        """Convertit une référence de cellule - version avec safe_cell"""
        try:
            # Enlever les $ pour le traitement
            ref_clean = ref.replace('$', '')
            
            # Gérer les références avec feuille
            sheet = current_sheet
            cell_part = ref_clean
            
            if '!' in ref_clean:
                sheet_part, cell_part = ref_clean.split('!', 1)
                sheet = sheet_part.strip("'")
            
            # Si c'est une plage
            if ':' in cell_part:
                start, end = cell_part.split(':')
                return self._convert_range_reference(start, end, sheet)
            
            # Référence simple
            match = re.match(r'^([A-Z]+)(\d+)$', cell_part, re.IGNORECASE)
            if match:
                col_str, row_str = match.groups()
                col = self.excel_col_to_num(col_str)
                row = int(row_str)
                
                # S'assurer que les indices sont valides
                if row < 1 or col < 1:
                    return "0"
                
                # Utiliser safe_cell pour un accès robuste
                if sheet == current_sheet:
                    return f"safe_cell(ws, {row-1}, {col-1})"
                else:
                    return f"safe_cell(sheets['{sheet}'], {row-1}, {col-1})"
            
            return "0"
        
        except Exception as e:
            return "0"
        
    def _convert_range_reference(self, start: str, end: str, sheet: str) -> str:
        """Convertit une plage Excel en slice Python - version sécurisée"""
        start_match = re.match(r'^([A-Z]+)(\d+)$', start)
        end_match = re.match(r'^([A-Z]+)(\d+)$', end)
        
        if start_match and end_match:
            start_col = self.excel_col_to_num(start_match.group(1))
            start_row = int(start_match.group(2))
            end_col = self.excel_col_to_num(end_match.group(1))
            end_row = int(end_match.group(2))
            
            # Ajuster les indices (Excel est 1-based, Python est 0-based)
            start_row -= 1
            start_col -= 1
            end_row -= 1
            end_col -= 1
            
            # S'assurer que les indices sont positifs
            if any(x < 0 for x in [start_row, start_col, end_row, end_col]):
                return f"# Invalid range (negative indices): {start}:{end}"
            
            return f"sheets['{sheet}'].iloc[{start_row}:{end_row+1}, {start_col}:{end_col+1}]"
        
        return f"# Invalid range: {start}:{end}"
    # Convertisseurs de fonctions
    def _convert_sum(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        """Convertit SUM - version simplifiée avec helpers"""
        if len(args) == 0:
            return "0"
        elif len(args) == 1:
            arg = args[0]
            # Vérifier si c'est une plage
            if ".iloc[" in arg and ":" in arg:
                return f"safe_sum_range({arg})"
            else:
                # C'est une cellule unique, déjà convertie par safe_cell
                return arg
        else:
            # Plusieurs arguments
            sum_parts = []
            for arg in args:
                if ".iloc[" in arg and ":" in arg:
                    # C'est une plage
                    sum_parts.append(f"safe_sum_range({arg})")
                else:
                    # C'est une cellule unique
                    sum_parts.append(arg)
            
            # Simple addition
            return f"({' + '.join(sum_parts)})"

    def _convert_average(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        """Convertit AVERAGE - version corrigée pour plages multiples"""
        if len(args) == 0:
            return "np.nan"
        elif len(args) == 1:
            arg = args[0]
            if ".iloc[" in arg and ":" in arg:
                return f"np.nanmean({arg}.values)"
            else:
                return f"({arg} if not pd.isna({arg}) else np.nan)"
        else:
            # Pour plusieurs arguments, il faut collecter toutes les valeurs
            values_parts = []
            for arg in args:
                if ".iloc[" in arg and ":" in arg:
                    values_parts.append(f"{arg}.values.flatten()")
                else:
                    values_parts.append(f"[{arg}]")
            
            return f"np.nanmean(np.concatenate([{', '.join(values_parts)}]))"

    def _convert_max(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        """Convertit MAX - version ultra simple"""
        if len(args) == 0:
            return "0"
        else:
            # MAX prend simplement le maximum
            return f"max([{', '.join(args)}])"

    def _convert_min(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        """Convertit MIN - version simple"""
        if len(args) == 0:
            return "0"
        else:
            return f"min([{', '.join(args)}])"
        
    def _convert_count(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) == 1:
            if ".iloc[" in args[0] and ":" in args[0]:
                return f"np.count_nonzero(~pd.isna({args[0]}.values))"
            else:
                return f"(0 if pd.isna({args[0]}) else 1)"
        else:
            return f"sum(not pd.isna(x) for x in [{', '.join(args)}])"
        
    def _convert_counta(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) == 1:
            return f"np.count_nonzero({args[0]}.values.flatten() != '')"
        else:
            return f"sum(x != '' and x is not None for x in [{', '.join(args)}])"
    
    def _convert_round(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) >= 2:
            return f"np.round({args[0]}, {args[1]})"
        else:
            return f"np.round({args[0]})"
    
    def _convert_int(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"int({args[0]})"
    
    def _convert_if(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) >= 3:
            return f"({args[1]} if {args[0]} else {args[2]})"
        elif len(args) == 2:
            return f"({args[1]} if {args[0]} else '')"
        else:
            return "# Invalid IF"
    
    def _convert_sumif(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) < 2:
            return "# SUMIF requires at least 2 arguments"
        
        range_arg = args[0]
        criteria = args[1]
        sum_range = args[2] if len(args) >= 3 else range_arg
        
        # Simplifier le critère
        if criteria.startswith('"') and criteria.endswith('"'):
            criteria_value = criteria[1:-1]
            return f"np.nansum({sum_range}.values[{range_arg}.values == '{criteria_value}'])"
        elif self._is_number(criteria):
            return f"np.nansum({sum_range}.values[{range_arg}.values == {criteria}])"
        else:
            # Critère plus complexe
            return f"np.nansum({sum_range}.values[{range_arg}.values == {criteria}])"
        
    def _convert_sumif_criteria(self, criteria: str, range_ref: str, current_sheet: str) -> str:
        """Convertit un critère SUMIF en condition Python"""
        criteria = criteria.strip()
        
        # Critères de comparaison
        if criteria.startswith(('>', '<', '=')):
            op_match = re.match(r'^([><=]+)(.*)$', criteria)
            if op_match:
                op, value = op_match.groups()
                if op == '=':
                    op = '=='
                
                # Si c'est un nombre
                if self._is_number(value):
                    return f"{range_ref}.values {op} {value}"
                else:
                    # C'est une chaîne
                    return f"{range_ref}.values {op} {self._parse_formula(value, current_sheet)}"
        
        # Wildcard patterns (* et ?)
        if '*' in criteria or '?' in criteria:
            # Convertir en regex
            pattern = criteria.replace('*', '.*').replace('?', '.')
            return f"{range_ref}.astype(str).str.match('{pattern}')"
        
        # Référence de cellule
        if self._is_cell_reference(criteria):
            ref_value = self._parse_formula(criteria, current_sheet)
            return f"{range_ref}.values == {ref_value}"
        
        # Valeur simple
        if self._is_number(criteria):
            return f"{range_ref}.values == {criteria}"
        else:
            # Chaîne
            criteria_clean = criteria.strip('"')
            return f"{range_ref}.values == '{criteria_clean}'"
    
    def _convert_iferror(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) >= 2:
            return f"(lambda: {args[1]} if pd.isna({args[0]}) or isinstance({args[0]}, Exception) else {args[0]})()"
        else:
            return "# Invalid IFERROR"
    
    def _convert_text(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"str({args[0]})"
    
    def _convert_concatenate(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"''.join(str(x) for x in [{', '.join(args)}])"
    
    def _convert_and(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"({' and '.join(args)})"
    
    def _convert_or(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"({' or '.join(args)})"
    
    def _convert_not(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"(not {args[0]})"
    
    def _convert_isblank(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"(pd.isna({args[0]}) or {args[0]} == '')"
    
    def _convert_date(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) >= 3:
            return f"pd.Timestamp({args[0]}, {args[1]}, {args[2]})"
        else:
            return "# Invalid DATE"
    
    def _convert_offset(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        """Convertit OFFSET - basé sur la logique R"""
        if len(args) < 3:
            return "# OFFSET requires at least 3 arguments"
        
        # Parser la référence de base
        base_ref = raw_args[0].strip()
        row_offset = args[1]
        col_offset = args[2]
        
        # Extraire les coordonnées de base
        if '!' in base_ref:
            sheet, cell = base_ref.split('!', 1)
            sheet = sheet.strip("'")
        else:
            sheet = current_sheet
            cell = base_ref
        
        match = re.match(r'^([A-Z]+)(\d+)$', cell)
        if match:
            base_col = self.excel_col_to_num(match.group(1))
            base_row = int(match.group(2))
            
            return f"sheets['{sheet}'].iloc[{base_row-1} + ({row_offset}), {base_col-1} + ({col_offset})]"
        
        return "# Invalid OFFSET reference"
    
    def _convert_indirect(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        # INDIRECT est complexe car il évalue dynamiquement
        return f"eval({args[0]})"  # Attention: utiliser avec précaution
    
    def _convert_vlookup(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) < 3:
            return "# VLOOKUP requires at least 3 arguments"
        
        lookup_value = args[0]
        table_array = args[1]
        col_index = args[2]
        range_lookup = args[3] if len(args) > 3 else "True"
        
        return f"vlookup({lookup_value}, {table_array}, {col_index}, {range_lookup})"
    
    def _convert_index(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) >= 2:
            return f"{args[0]}.iloc[{args[1]}-1" + (f", {args[2]}-1]" if len(args) > 2 else "]")
        return "# Invalid INDEX"
    
    def _convert_match(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) >= 2:
            lookup_value = args[0]
            lookup_array = args[1]
            match_type = args[2] if len(args) > 2 else "1"
            
            return f"match_index({lookup_value}, {lookup_array}, {match_type})"
        return "# Invalid MATCH"
    
    def _convert_len(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"len(str({args[0]}))"
    
    def _convert_trim(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"str({args[0]}).strip()"
    
    def _convert_upper(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"str({args[0]}).upper()"
    
    def _convert_lower(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"str({args[0]}).lower()"
    
    def _convert_mid(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) >= 3:
            text = args[0]
            start = args[1]
            length = args[2]
            return f"str({text})[{start}-1:{start}-1+{length}]"
        return "# Invalid MID"
    
    def _convert_substitute(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) >= 3:
            text = args[0]
            old = args[1]
            new = args[2]
            instance = args[3] if len(args) > 3 else None
            
            if instance:
                return f"substitute({text}, {old}, {new}, {instance})"
            else:
                return f"str({text}).replace({old}, {new})"
        return "# Invalid SUBSTITUTE"
    
    def _convert_abs(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"abs({args[0]})"
    
    def _convert_sqrt(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        return f"np.sqrt({args[0]})"
    
    def _convert_power(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) >= 2:
            return f"pow({args[0]}, {args[1]})"
        return "# Invalid POWER"
    
    def _convert_mod(self, args: List[str], current_sheet: str, raw_args: List[str]) -> str:
        if len(args) >= 2:
            return f"({args[0]} % {args[1]})"
        return "# Invalid MOD"

    def _generate_python_script(self, formulas: List[FormulaCell], original_file: str) -> str:
        """Génère un script Python pour appliquer les formules"""
        script_name = original_file.replace('.xlsx', '_formulas.py')
        
        with open(script_name, 'w', encoding='utf-8') as f:
            f.write("""# Auto-generated Excel formula script
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime

# Helper functions
def safe_get_numeric(value, default=0):
    \"\"\"Convertit une valeur en nombre de manière sûre\"\"\"
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        # Si c'est une formule Excel, retourner default
        if value.startswith('='):
            return default
        # Essayer de convertir
        try:
            # Gérer les virgules comme séparateurs décimaux
            cleaned = value.replace(',', '.').replace(' ', '')
            return float(cleaned) if cleaned else default
        except:
            return default
    # Pour tout autre type
    try:
        return float(value)
    except:
        return default

def safe_sum_range(df_range):
    \"\"\"Somme sûre d'une plage de DataFrame\"\"\"
    if df_range is None:
        return 0
    try:
        # Aplatir et convertir en numérique
        values = df_range.values.flatten()
        numeric_values = [safe_get_numeric(v) for v in values]
        return sum(numeric_values)
    except:
        return 0

def safe_cell(df, row, col, default=0):
    \"\"\"Accès sûr à une cellule\"\"\"
    try:
        if row < 0 or col < 0:
            return default
        if row >= len(df) or col >= len(df.columns):
            return default
        value = df.iloc[row, col]
        return safe_get_numeric(value, default)
    except:
        return default

def vlookup(lookup_value, table_array, col_index, range_lookup=True):
    \"\"\"VLOOKUP implementation\"\"\"
    try:
        if range_lookup:
            # Approximate match
            idx = table_array.iloc[:, 0].searchsorted(lookup_value)
            if idx > 0:
                idx -= 1
        else:
            # Exact match
            mask = table_array.iloc[:, 0] == lookup_value
            if mask.any():
                idx = mask.idxmax()
            else:
                return np.nan
        
        return table_array.iloc[idx, col_index - 1]
    except:
        return np.nan

def match_index(lookup_value, lookup_array, match_type=1):
    \"\"\"MATCH implementation\"\"\"
    try:
        if match_type == 0:
            # Exact match
            return (lookup_array == lookup_value).idxmax() + 1
        elif match_type == 1:
            # Less than or equal
            return lookup_array[lookup_array <= lookup_value].idxmax() + 1
        else:
            # Greater than or equal
            return lookup_array[lookup_array >= lookup_value].idxmin() + 1
    except:
        return np.nan

def substitute(text, old, new, instance=None):
    \"\"\"SUBSTITUTE implementation\"\"\"
    text = str(text)
    if instance is None:
        return text.replace(old, new)
    else:
        parts = text.split(old)
        if len(parts) > instance:
            return old.join(parts[:instance]) + new + old.join(parts[instance:])
        return text

def apply_formulas(excel_file):
    \"\"\"Apply Excel formulas to the workbook\"\"\"
    # Load workbook
    wb = openpyxl.load_workbook(excel_file, data_only=False)
    
    # Load all sheets into DataFrames
    sheets = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        sheets[sheet_name] = pd.DataFrame(data)
    
    # Apply formulas
""")
            
            # Grouper par feuille
            from collections import defaultdict
            formulas_by_sheet = defaultdict(list)
            for formula in formulas:
                if formula.python_code and not formula.error:
                    formulas_by_sheet[formula.sheet].append(formula)
            
            # Générer le code pour chaque feuille
            for sheet_name, sheet_formulas in formulas_by_sheet.items():
                f.write(f"\n    # Sheet: {sheet_name}\n")
                f.write(f"    ws = sheets['{sheet_name}']\n")
                
                for formula in sheet_formulas:
                    f.write(f"    # {formula.address}: {formula.formula}\n")
                    f.write(f"    try:\n")
                    f.write(f"        result = {formula.python_code}\n")
                    f.write(f"        wb[\"{sheet_name}\"][\"{formula.address}\"].value = result\n")
                    f.write(f"        ws.iloc[{formula.row-1}, {formula.col-1}] = result\n")
                    f.write(f"    except Exception as e:\n")
                    f.write(f"        print(f'Error in {formula.address}: {{e}}')\n")
                    f.write(f"    \n")
            
            f.write("""
    # Save workbook
    output_file = excel_file.replace('.xlsx', '_calculated.xlsx')
    wb.save(output_file)
    wb.close()
    print(f"Formulas applied. Output saved to: {output_file}")
    return sheets

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        apply_formulas(sys.argv[1])
    else:
        print("Usage: python script.py <excel_file>")
""")
        
        logger.info(f"Generated script: {script_name}")
        return script_name

    def apply_formulas_to_workbook(self, workbook: openpyxl.Workbook, 
                                formulas: List[FormulaCell]) -> openpyxl.Workbook:
        """Applique les formules au workbook avec gestion robuste des types"""
        
        # Charger toutes les feuilles en DataFrames avec conversion appropriée
        sheets = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = []
            
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    value = cell.value
                    
                    # Si c'est une formule, essayer d'obtenir la valeur calculée
                    if isinstance(value, str) and value.startswith('='):
                        # Pour les formules, mettre 0 par défaut
                        row_data.append(0)
                    elif value is None:
                        row_data.append(0)  # Remplacer None par 0
                    elif isinstance(value, (int, float)):
                        row_data.append(value)
                    elif isinstance(value, str):
                        # Essayer de convertir en nombre
                        try:
                            # Gérer les formats européens (virgule comme séparateur décimal)
                            cleaned = value.replace(' ', '').replace(',', '.')
                            if cleaned.replace('.', '').replace('-', '').replace('+', '').isdigit():
                                row_data.append(float(cleaned))
                            else:
                                # Si ce n'est pas un nombre, mettre 0
                                row_data.append(0)
                        except:
                            row_data.append(0)
                    else:
                        # Pour tout autre type, convertir ou mettre 0
                        try:
                            row_data.append(float(value))
                        except:
                            row_data.append(0)
                
                data.append(row_data)
            
            # Créer le DataFrame
            sheets[sheet_name] = pd.DataFrame(data) if data else pd.DataFrame()
        
        # Créer l'environnement d'exécution avec helpers inline
        exec_globals = {
            'sheets': sheets,
            'np': np,
            'pd': pd,
            'datetime': datetime,
            'vlookup': ExcelFormulaParser._vlookup_impl,
            'match_index': ExcelFormulaParser._match_index_impl,
            'substitute': ExcelFormulaParser._substitute_impl,
            # Ajouter des fonctions helper inline
            'safe_value': lambda x: 0 if (x is None or pd.isna(x) or (isinstance(x, str) and x.startswith('='))) else x,
            'safe_div': lambda a, b: a / b if b != 0 else 0,
        }

        # Trier les formules par dépendances
        sorted_formulas = self._topological_sort(formulas)
        
        # Appliquer chaque formule
        success_count = 0
        error_count = 0
        errors_list = []
        
        for formula in formulas:
            if formula.python_code and not formula.error:
                try:
                    # Créer un environnement local avec la feuille courante
                    exec_locals = {'ws': sheets[formula.sheet]}
                    
                    # Évaluer la formule
                    result = eval(formula.python_code, exec_globals, exec_locals)
                    
                    # IMPORTANT : Vérifier et convertir le résultat avant de l'écrire
    
                    # Si le résultat est None, le remplacer par 0 ou une chaîne vide
                    if result is None:
                        logger.warning(f"Formula {formula.sheet}!{formula.address} returned None")
                        result = 0  # ou "" selon le contexte

                    # Gérer les différents types de résultats
                    if isinstance(result, pd.DataFrame):
                        if result.size == 0:
                            result = 0
                        elif result.size == 1:
                            result = result.iloc[0, 0]
                        else:
                            # Pour un DataFrame multi-valeurs, prendre la première valeur
                            result = result.iloc[0, 0] if result.shape[1] > 0 else 0
                            
                    elif isinstance(result, pd.Series):
                        if len(result) == 0:
                            result = 0
                        elif len(result) == 1:
                            result = result.iloc[0]
                        else:
                            result = result.iloc[0]
                            
                    elif isinstance(result, np.ndarray):
                        if result.size == 0:
                            result = 0
                        elif result.size == 1:
                            result = result.item()
                        else:
                            result = result.flat[0]
                    
                    # S'assurer que le résultat n'est pas None après conversion
                    if result is None:
                        result = 0

                    # Pour les résultats NaN, les convertir en 0 ou laisser vide
                    if isinstance(result, (int, float)) and np.isnan(result):
                        result = 0  # ou None selon le besoin
                    
                    # Vérifier le type avant d'écrire
                    if isinstance(result, (list, tuple, dict, set)):
                        logger.error(f"Formula {formula.sheet}!{formula.address} returned non-scalar type: {type(result)}")
                        result = str(result)  # Convertir en string pour éviter l'erreur

                    # Mettre à jour le workbook
                    sheet = workbook[formula.sheet]
                    cell = sheet.cell(row=formula.row, column=formula.col)

                    # IMPORTANT : S'assurer que la valeur est compatible avec openpyxl
                    # openpyxl accepte : int, float, str, bool, datetime, None
                    if not isinstance(result, (int, float, str, bool, type(None))):
                        result = str(result)
    
                    cell.value = result

                    # Mettre à jour le DataFrame aussi pour les calculs suivants
                    sheets[formula.sheet].iloc[formula.row-1, formula.col-1] = result
                    
                    formula.value = result
                    success_count += 1
                    
                except Exception as e:
                    formula.error = f"Erreur de calcul: {str(e)}"
                    error_count += 1
                    errors_list.append({
                        'cell': f"{formula.sheet}!{formula.address}",
                        'formula': formula.formula,
                        'error': str(e),
                        'python_code': formula.python_code
                    })
                    logger.error(f"Erreur dans {formula.sheet}!{formula.address}: {str(e)}")
                    logger.debug(f"Code Python: {formula.python_code}")
        # Stocker les erreurs dans session_state pour l'affichage
        import streamlit as st
        if 'formula_errors' not in st.session_state:
            st.session_state.formula_errors = []
        st.session_state.formula_errors = errors_list
        
        logger.info(f"Formules appliquées: {success_count} succès, {error_count} erreurs")
        return workbook

    # Implémentations des fonctions helper
    @staticmethod
    def _vlookup_impl(lookup_value, table_array, col_index, range_lookup=True):
        """Implémentation de VLOOKUP"""
        try:
            if range_lookup:
                idx = table_array.iloc[:, 0].searchsorted(lookup_value)
                if idx > 0:
                    idx -= 1
            else:
                mask = table_array.iloc[:, 0] == lookup_value
                if mask.any():
                    idx = mask.idxmax()
                else:
                    return np.nan
            
            return table_array.iloc[idx, col_index - 1]
        except:
            return np.nan
    
    @staticmethod
    def _match_index_impl(lookup_value, lookup_array, match_type=1):
        """Implémentation de MATCH"""
        try:
            if match_type == 0:
                return (lookup_array == lookup_value).idxmax() + 1
            elif match_type == 1:
                return lookup_array[lookup_array <= lookup_value].idxmax() + 1
            else:
                return lookup_array[lookup_array >= lookup_value].idxmin() + 1
        except:
            return np.nan
    
    @staticmethod
    def _substitute_impl(text, old, new, instance=None):
        """Implémentation de SUBSTITUTE"""
        text = str(text)
        if instance is None:
            return text.replace(old, new)
        else:
            parts = text.split(old)
            if len(parts) > instance:
                return old.join(parts[:instance]) + new + old.join(parts[instance:])
            return text