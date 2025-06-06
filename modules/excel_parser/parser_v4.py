import logging
from copy import copy
from dataclasses import dataclass
from typing import Optional
import tempfile
import os

import openpyxl
import formulas

logger = logging.getLogger(__name__)

@dataclass
class ParserConfig:
    handle_circular_references: bool = False

class SimpleExcelFormulaParser:
    def __init__(self, config: Optional[ParserConfig] = None):
        self.config = config or ParserConfig()

    def parse_and_apply(self, source: openpyxl.Workbook):
        """
        Parses formulas in the source workbook and applies them to the target workbook.
        Returns a copy of the source workbook with formulas parsed and applied.
        """

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
            source.save(tmp_path)
            logger.debug(f"Temporary file created at {tmp_path}")

        try:
            # Calculate all formulas using formulas library
            xl_model = formulas.ExcelModel().loads(tmp_path).finish()
            calculated_values = xl_model.calculate()

            # Create completely new workbook (copy)
            new_wb = openpyxl.Workbook()

            # Remove default sheet
            new_wb.remove(new_wb.active)

            # Copy each worksheet
            for sheet_name in source.sheetnames:
                original_ws = source[sheet_name]
                new_ws = new_wb.create_sheet(sheet_name)

                # Copy all cells
                for row in original_ws.iter_rows():
                    for cell in row:
                        new_cell = new_ws.cell(row=cell.row, column=cell.column)

                        # If it's a formula, use calculated value
                        if cell.data_type == 'f':
                            cell_address = f"{sheet_name}!{cell.coordinate}"
                            if cell_address in calculated_values:
                                new_cell.value = calculated_values[cell_address]
                            else:
                                # If calculation failed, keep original formula
                                new_cell.value = cell.value
                        else:
                            # Regular value - copy as is
                            new_cell.value = cell.value

                        # Copy cell formatting
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = cell.number_format
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)

            return new_wb

        finally:
            os.unlink(tmp_path)

