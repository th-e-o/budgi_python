# modules/budget_mapper.py - VERSION CORRIGÉE
import pandas as pd
from typing import List, Dict, Optional, Tuple, Set
import logging
import re
from difflib import SequenceMatcher
import asyncio
from collections import defaultdict
from datetime import datetime
import json
import openpyxl

logger = logging.getLogger(__name__)

class BudgetMapper:
    """Module optimisé pour mapper les entrées budgétaires aux cellules Excel"""
    
    def __init__(self, llm_client):
        self.llm_client = llm_client
        self.mapping_cache = {}
        self.batch_size = 5
        self.max_tags_per_batch = 50
        
        # Poids pour le scoring
        self.weights = {
            'description_match': 0.22,
            'axe_match': 0.22,
            'montant_match': 0,
            'date_match': 0.22,
            'nature_match': 0.22,
            'unite_match': 0,
            'source_context': 0.12
        }
    
    async def map_entries_to_cells(self, entries: List[Dict], tags: List[Dict], 
                                  progress_callback=None) -> List[Dict]:
        """Mappe les entrées budgétaires aux cellules Excel de manière itérative"""
        if not entries or not tags:
            return []
        
        logger.info(f"Début du mapping: {len(entries)} entrées vers {len(tags)} tags")
        
        # Enrichir les tags avec des métadonnées
        enriched_tags = self._enrich_tags(tags)
        
        # Créer des index pour accélérer la recherche
        tag_indexes = self._build_comprehensive_indexes(enriched_tags)
        
        # Résultats du mapping
        all_mappings = []
        
        # Traiter les entrées par batch
        total_batches = (len(entries) - 1) // self.batch_size + 1
        
        for batch_idx in range(0, len(entries), self.batch_size):
            batch_entries = entries[batch_idx:batch_idx + self.batch_size]
            current_batch = batch_idx // self.batch_size + 1
            
            logger.info(f"Traitement du batch {current_batch}/{total_batches}")
            
            if progress_callback:
                progress = (batch_idx / len(entries)) * 100
                progress_callback(progress, f"Traitement batch {current_batch}/{total_batches}")
            
            batch_mappings = []
            
            for entry in batch_entries:
                # Enrichir l'entrée avec des métadonnées
                enriched_entry = self._enrich_entry(entry)
                
                # Utiliser le cache si disponible
                cache_key = self._get_cache_key(entry)
                if cache_key in self.mapping_cache:
                    batch_mappings.append(self.mapping_cache[cache_key])
                    continue
                
                # Trouver les tags candidats
                candidate_tags = self._find_candidate_tags_advanced(
                    enriched_entry, tag_indexes, enriched_tags
                )
                
                if not candidate_tags:
                    logger.warning(f"Aucun tag candidat pour: {entry.get('Description', '')}")
                    # Créer un mapping vide pour tracking
                    empty_mapping = self._create_empty_mapping(entry)
                    batch_mappings.append(empty_mapping)
                    continue
                
                # Décision de mapping
                if len(candidate_tags) == 1 and candidate_tags[0]['score'] > 0.85:
                    # Mapping évident
                    best_tag = candidate_tags[0]['tag']
                    mapping = self._create_detailed_mapping(entry, best_tag, candidate_tags[0]['score'])
                    batch_mappings.append(mapping)
                    self.mapping_cache[cache_key] = mapping
                elif candidate_tags[0]['score'] > 0.7:
                    # Probablement bon, prendre le meilleur
                    best_tag = candidate_tags[0]['tag']
                    mapping = self._create_detailed_mapping(entry, best_tag, candidate_tags[0]['score'])
                    batch_mappings.append(mapping)
                    self.mapping_cache[cache_key] = mapping
                else:
                    # Cas ambigu, utiliser le LLM
                    best_mapping = await self._llm_select_best_tag_advanced(
                        enriched_entry, candidate_tags[:10]  # Top 10 candidats
                    )
                    if best_mapping:
                        batch_mappings.append(best_mapping)
                        self.mapping_cache[cache_key] = best_mapping
                    else:
                        # Fallback sur le meilleur score
                        best_tag = candidate_tags[0]['tag']
                        mapping = self._create_detailed_mapping(entry, best_tag, candidate_tags[0]['score'])
                        batch_mappings.append(mapping)
            
            all_mappings.extend(batch_mappings)
            
            # Pause entre les batches pour éviter le rate limiting
            if batch_idx + self.batch_size < len(entries):
                await asyncio.sleep(1.5)
        
        logger.info(f"Mapping terminé: {len(all_mappings)} mappings créés")
        return all_mappings
    
    def apply_mapping_to_excel(self, workbook: openpyxl.Workbook, mapping: List[Dict], 
                              entries_data: pd.DataFrame) -> Tuple[int, List[str]]:
        """
        Applique le mapping au workbook Excel
        CORRIGÉE : utilise correctement openpyxl pour écrire les valeurs
        """
        success_count = 0
        errors = []
        warnings = []
        modified_cells = []  # Pour tracer les modifications
        
        # Créer un index pour lookup rapide
        mapping_index = {}
        for m in mapping:
            if not m.get('not_mapped', False) and m.get('cellule'):
                key = (m.get('Axe'), m.get('Description'), m.get('Montant'))
                mapping_index[key] = m
        
        logger.info(f"Application du mapping: {len(mapping_index)} entrées à mapper")
        
        for idx, entry in entries_data.iterrows():
            try:
                key = (entry.get('Axe'), entry.get('Description'), entry.get('Montant'))
                if key not in mapping_index:
                    continue
                
                mapping_entry = mapping_index[key]
                
                # Vérifier le score de confiance
                confidence = mapping_entry.get('confidence_score', 0)
                if confidence < 0.5:
                    warnings.append(
                        f"⚠️ Mapping peu fiable ({confidence:.1%}) "
                        f"pour: {entry.get('Description', '')[:50]}..."
                    )
                
                sheet_name = mapping_entry.get('sheet_name', 'Sheet1')
                cell_address = mapping_entry.get('cellule', '').strip()
                montant = entry.get('Montant', 0)
                
                # Vérification de l'adresse
                if not cell_address:
                    errors.append(f"Pas d'adresse de cellule pour: {entry.get('Description')}")
                    continue
                
                # Vérifier que la feuille existe
                if sheet_name not in workbook.sheetnames:
                    errors.append(f"Feuille '{sheet_name}' non trouvée dans le workbook")
                    continue
                
                # Obtenir la feuille
                sheet = workbook[sheet_name]
                
                # Écrire directement avec l'adresse de cellule
                try:
                    # Méthode 1 : Utiliser directement l'adresse
                    cell = sheet[cell_address]
                    
                    # S'assurer que le montant est un nombre
                    if isinstance(montant, (int, float)):
                        cell.value = montant
                    else:
                        # Essayer de convertir
                        try:
                            cell.value = float(montant)
                        except:
                            cell.value = 0
                            warnings.append(f"Montant invalide converti en 0 pour {cell_address}")
                    
                    success_count += 1
                    modified_cells.append({
                        'sheet': sheet_name,
                        'cell': cell_address,
                        'value': cell.value,
                        'description': entry.get('Description', '')[:50]
                    })
                    
                    logger.info(
                        f"✓ Écrit {cell.value} dans {sheet_name}!{cell_address} "
                        f"(confiance: {confidence:.1%})"
                    )
                    
                except Exception as e:
                    # Si erreur avec l'adresse directe, essayer avec parsing
                    try:
                        row, col = self._parse_cell_address(cell_address)
                        sheet.cell(row=row, column=col, value=montant)
                        
                        success_count += 1
                        modified_cells.append({
                            'sheet': sheet_name,
                            'cell': cell_address,
                            'value': montant,
                            'description': entry.get('Description', '')[:50]
                        })
                        
                        logger.info(f"✓ Écrit {montant} dans {sheet_name}!{cell_address} (méthode alternative)")
                        
                    except Exception as e2:
                        errors.append(f"Erreur écriture {sheet_name}!{cell_address}: {str(e2)}")
                        logger.error(f"Erreur écriture cellule: {str(e2)}")
                    
            except Exception as e:
                errors.append(f"Erreur ligne {idx}: {str(e)}")
                logger.error(f"Erreur mapping ligne {idx}: {str(e)}")
        
        # Log détaillé des résultats
        logger.info(f"Mapping appliqué: {success_count} succès, {len(errors)} erreurs, {len(warnings)} warnings")
        
        # Afficher un résumé des modifications
        if modified_cells:
            logger.info("Cellules modifiées:")
            for mod in modified_cells[:10]:  # Afficher les 10 premières
                logger.info(f"  - {mod['sheet']}!{mod['cell']} = {mod['value']} ({mod['description']})")
            if len(modified_cells) > 10:
                logger.info(f"  ... et {len(modified_cells) - 10} autres modifications")
        
        # Ajouter les warnings aux erreurs pour le retour
        all_issues = errors + warnings
        
        # Retourner aussi les cellules modifiées pour traçabilité
        return success_count, all_issues, modified_cells
    
    def validate_and_prepare_mapping(self, mapping: List[Dict], workbook: openpyxl.Workbook) -> Tuple[List[Dict], List[str]]:
        """
        Valide le mapping avant application et prépare les données
        Retourne le mapping validé et les problèmes détectés
        """
        validated_mapping = []
        issues = []
        
        # Obtenir la liste des feuilles disponibles
        available_sheets = workbook.sheetnames
        
        for idx, entry in enumerate(mapping):
            # Copier l'entrée pour ne pas modifier l'original
            validated_entry = entry.copy()
            
            # Vérifications et corrections
            sheet_name = validated_entry.get('sheet_name', '').strip()
            cell_address = validated_entry.get('cellule', '').strip()
            
            # 1. Vérifier la feuille
            if not sheet_name:
                issues.append(f"Entrée {idx}: Pas de feuille spécifiée")
                continue
            
            if sheet_name not in available_sheets:
                # Chercher une feuille similaire
                similar = self._find_similar_sheet(sheet_name, available_sheets)
                if similar:
                    validated_entry['sheet_name'] = similar
                    issues.append(f"Entrée {idx}: Feuille '{sheet_name}' remplacée par '{similar}'")
                else:
                    issues.append(f"Entrée {idx}: Feuille '{sheet_name}' introuvable")
                    continue
            
            # 2. Vérifier et nettoyer l'adresse de cellule
            if not cell_address:
                issues.append(f"Entrée {idx}: Pas d'adresse de cellule")
                continue
            
            # Nettoyer l'adresse (enlever espaces, normaliser)
            cleaned_address = cell_address.upper().strip()
            if not re.match(r'^[A-Z]+\d+$', cleaned_address):
                issues.append(f"Entrée {idx}: Adresse invalide '{cell_address}'")
                continue
            
            validated_entry['cellule'] = cleaned_address
            
            # 3. Vérifier le montant
            montant = validated_entry.get('Montant')
            if montant is None:
                validated_entry['Montant'] = 0
            elif not isinstance(montant, (int, float)):
                try:
                    validated_entry['Montant'] = float(montant)
                except:
                    validated_entry['Montant'] = 0
                    issues.append(f"Entrée {idx}: Montant invalide converti en 0")
            
            validated_mapping.append(validated_entry)
        
        logger.info(f"Validation terminée: {len(validated_mapping)} entrées valides, {len(issues)} problèmes")
        return validated_mapping, issues
    
    def _find_similar_sheet(self, target: str, available: List[str]) -> Optional[str]:
        """Trouve une feuille similaire dans la liste"""
        target_lower = target.lower()
        
        # Recherche exacte (insensible à la casse)
        for sheet in available:
            if sheet.lower() == target_lower:
                return sheet
        
        # Recherche par similarité
        best_match = None
        best_score = 0
        
        for sheet in available:
            score = SequenceMatcher(None, target_lower, sheet.lower()).ratio()
            if score > best_score and score > 0.8:  # Seuil de 80%
                best_score = score
                best_match = sheet
        
        return best_match
    
    def _parse_cell_address(self, address: str) -> Tuple[int, int]:
        """Parse une adresse de cellule Excel (ex: 'A1' -> (1, 1))"""
        match = re.match(r'^([A-Z]+)(\d+)$', address.upper())
        if not match:
            raise ValueError(f"Adresse invalide: {address}")
        
        col_str, row_str = match.groups()
        
        # Convertir la colonne en numéro (A=1, B=2, ..., AA=27, etc.)
        col = 0
        for char in col_str:
            col = col * 26 + (ord(char) - ord('A') + 1)
        
        row = int(row_str)
        
        # Vérifier les limites Excel
        if row < 1 or row > 1048576:  # Limite Excel
            raise ValueError(f"Numéro de ligne invalide: {row}")
        if col < 1 or col > 16384:  # Limite Excel (XFD)
            raise ValueError(f"Numéro de colonne invalide: {col}")
        
        return row, col
    
    def create_mapping_summary(self, mapping: List[Dict], modified_cells: List[Dict]) -> str:
        """Crée un résumé détaillé du mapping appliqué"""
        summary_lines = []
        
        # Statistiques générales
        total = len(mapping)
        applied = len(modified_cells)
        
        summary_lines.append(f"📊 RÉSUMÉ DU MAPPING")
        summary_lines.append(f"{'='*40}")
        summary_lines.append(f"Total d'entrées : {total}")
        summary_lines.append(f"Cellules modifiées : {applied}")
        summary_lines.append(f"Taux de succès : {(applied/total*100):.1f}%")
        summary_lines.append("")
        
        # Répartition par feuille
        by_sheet = defaultdict(int)
        for cell in modified_cells:
            by_sheet[cell['sheet']] += 1
        
        summary_lines.append("📋 PAR FEUILLE:")
        for sheet, count in sorted(by_sheet.items()):
            summary_lines.append(f"  - {sheet}: {count} cellules")
        summary_lines.append("")
        
        # Montants totaux par feuille
        by_sheet_amount = defaultdict(float)
        for cell in modified_cells:
            by_sheet_amount[cell['sheet']] += float(cell.get('value', 0))
        
        summary_lines.append("💰 MONTANTS PAR FEUILLE:")
        for sheet, total in sorted(by_sheet_amount.items()):
            summary_lines.append(f"  - {sheet}: {total:,.2f} €")
        
        return "\n".join(summary_lines)
    
    # Les autres méthodes restent identiques...
    # (Je ne les répète pas pour économiser de l'espace, mais elles sont incluses)
    
    def _enrich_entry(self, entry: Dict) -> Dict:
        """Enrichit une entrée avec des métadonnées supplémentaires"""
        enriched = entry.copy()
        
        # Extraire des informations supplémentaires
        text = f"{entry.get('Axe', '')} {entry.get('Description', '')} {entry.get('SourcePhrase', '')}"
        
        # Années mentionnées
        enriched['years'] = list(set(re.findall(r'\b20\d{2}\b', text)))
        
        # Mots-clés importants
        keywords = []
        for word in re.findall(r'\b\w+\b', text.lower()):
            if len(word) > 4 and word not in self._get_stop_words():
                keywords.append(word)
        enriched['keywords'] = list(set(keywords))
        
        # Type de montant (si applicable)
        montant = entry.get('Montant', 0)
        if montant:
            if montant >= 1_000_000:
                enriched['montant_category'] = 'millions'
            elif montant >= 1_000:
                enriched['montant_category'] = 'milliers'
            else:
                enriched['montant_category'] = 'unites'
        
        # Nature normalisée
        nature = str(entry.get('Nature', '')).lower()
        if 'recette' in nature:
            enriched['nature_type'] = 'recette'
        elif 'depense' in nature or 'dépense' in nature:
            enriched['nature_type'] = 'depense'
        else:
            enriched['nature_type'] = 'autre'
        
        return enriched
    
    def _enrich_tags(self, tags: List[Dict]) -> List[Dict]:
        """Enrichit les tags avec des métadonnées"""
        enriched = []
        
        for tag in tags:
            enriched_tag = tag.copy()
            
            # Extraire toutes les informations des labels
            all_text = ' '.join(str(label) for label in tag.get('labels', []))
            
            # Années
            enriched_tag['years'] = list(set(re.findall(r'\b20\d{2}\b', all_text)))
            
            # Mots-clés
            keywords = []
            for word in re.findall(r'\b\w+\b', all_text.lower()):
                if len(word) > 4 and word not in self._get_stop_words():
                    keywords.append(word)
            enriched_tag['keywords'] = list(set(keywords))
            
            # Type de flux détecté
            if any(word in all_text.lower() for word in ['arrivée', 'entrant', 'recrutement']):
                enriched_tag['flux_type'] = 'entree'
            elif any(word in all_text.lower() for word in ['départ', 'sortant', 'sortie']):
                enriched_tag['flux_type'] = 'sortie'
            else:
                enriched_tag['flux_type'] = None
            
            # Catégorie détectée
            if 'effectif' in all_text.lower():
                enriched_tag['category'] = 'effectifs'
            elif 'emploi' in all_text.lower():
                enriched_tag['category'] = 'emplois'
            elif 'budget' in all_text.lower():
                enriched_tag['category'] = 'budget'
            else:
                enriched_tag['category'] = 'autre'
            
            enriched.append(enriched_tag)
        
        return enriched
    
    def _build_comprehensive_indexes(self, tags: List[Dict]) -> Dict:
        """Construit plusieurs index pour recherche multi-critères"""
        indexes = {
            'by_keyword': defaultdict(list),
            'by_year': defaultdict(list),
            'by_sheet': defaultdict(list),
            'by_flux_type': defaultdict(list),
            'by_category': defaultdict(list),
            'by_cell_pattern': defaultdict(list)
        }
        
        for tag in tags:
            # Index par mots-clés
            for keyword in tag.get('keywords', []):
                indexes['by_keyword'][keyword].append(tag)
            
            # Index par année
            for year in tag.get('years', []):
                indexes['by_year'][year].append(tag)
            
            # Index par feuille
            sheet = tag.get('sheet_name', '').lower()
            indexes['by_sheet'][sheet].append(tag)
            
            # Index par type de flux
            flux_type = tag.get('flux_type')
            if flux_type:
                indexes['by_flux_type'][flux_type].append(tag)
            
            # Index par catégorie
            category = tag.get('category', 'autre')
            indexes['by_category'][category].append(tag)
            
            # Index par pattern de cellule (colonne)
            cell_addr = tag.get('cell_address', '')
            if cell_addr:
                col_match = re.match(r'([A-Z]+)', cell_addr)
                if col_match:
                    col = col_match.group(1)
                    indexes['by_cell_pattern'][col].append(tag)
        
        return indexes
    
    def _find_candidate_tags_advanced(self, entry: Dict, indexes: Dict, 
                                    all_tags: List[Dict]) -> List[Dict]:
        """Trouve les tags candidats avec scoring avancé"""
        candidates = {}  # tag_id -> score
        
        # 1. Recherche par mots-clés
        for keyword in entry.get('keywords', []):
            if keyword in indexes['by_keyword']:
                for tag in indexes['by_keyword'][keyword]:
                    tag_id = tag.get('id')
                    if tag_id not in candidates:
                        candidates[tag_id] = {'tag': tag, 'score': 0, 'matches': []}
                    candidates[tag_id]['score'] += self.weights['description_match'] / len(entry.get('keywords', [1]))
                    candidates[tag_id]['matches'].append(f"keyword:{keyword}")
        
        # 2. Recherche par année
        for year in entry.get('years', []):
            if year in indexes['by_year']:
                for tag in indexes['by_year'][year]:
                    tag_id = tag.get('id')
                    if tag_id not in candidates:
                        candidates[tag_id] = {'tag': tag, 'score': 0, 'matches': []}
                    candidates[tag_id]['score'] += self.weights['date_match']
                    candidates[tag_id]['matches'].append(f"year:{year}")
        
        # 3. Bonus pour correspondance de nature/flux
        nature_type = entry.get('nature_type')
        if nature_type == 'recette' and nature_type in ['entree']:
            flux_candidates = indexes['by_flux_type'].get('entree', [])
        elif nature_type == 'depense' and nature_type in ['sortie']:
            flux_candidates = indexes['by_flux_type'].get('sortie', [])
        else:
            flux_candidates = []
        
        for tag in flux_candidates:
            tag_id = tag.get('id')
            if tag_id in candidates:
                candidates[tag_id]['score'] += self.weights['nature_match']
                candidates[tag_id]['matches'].append(f"flux_type:{tag.get('flux_type')}")
        
        # 4. Recherche par similarité textuelle pour tous les tags si peu de candidats
        if len(candidates) < 5:
            # Prendre un échantillon plus large
            sample_tags = all_tags[:200] if len(all_tags) > 200 else all_tags
            
            for tag in sample_tags:
                tag_id = tag.get('id')
                if tag_id not in candidates:
                    score = self._calculate_comprehensive_similarity(entry, tag)
                    if score > 0.3:
                        candidates[tag_id] = {
                            'tag': tag, 
                            'score': score, 
                            'matches': ['similarity']
                        }
        
        # 5. Calcul final des scores avec tous les critères
        final_candidates = []
        for tag_id, candidate_info in candidates.items():
            tag = candidate_info['tag']
            
            # Recalculer le score complet
            full_score = self._calculate_comprehensive_similarity(entry, tag)
            
            # Combiner avec le score des index
            combined_score = (candidate_info['score'] + full_score) / 2
            
            final_candidates.append({
                'tag': tag,
                'score': combined_score,
                'matches': candidate_info['matches'],
                'details': {
                    'index_score': candidate_info['score'],
                    'similarity_score': full_score
                }
            })
        
        # Trier par score décroissant
        final_candidates.sort(key=lambda x: x['score'], reverse=True)
        
        return final_candidates[:self.max_tags_per_batch]
    
    def _calculate_comprehensive_similarity(self, entry: Dict, tag: Dict) -> float:
        """Calcule un score de similarité complet entre une entrée et un tag"""
        scores = {}
        
        # 1. Description et Axe
        entry_desc = f"{entry.get('Axe', '')} {entry.get('Description', '')}".lower()
        tag_labels = ' '.join(str(label).lower() for label in tag.get('labels', []))
        
        # Similarité textuelle
        desc_similarity = SequenceMatcher(None, entry_desc, tag_labels).ratio()
        scores['description'] = desc_similarity * self.weights['description_match']
        
        # Bonus pour l'axe spécifique
        axe = str(entry.get('Axe', '')).lower()
        if axe and axe in tag_labels:
            scores['axe'] = self.weights['axe_match']
        else:
            scores['axe'] = 0
        
        # 2. Montant
        if 'Montant' in entry:
            montant_str = str(entry['Montant'])
            # Vérifier la présence exacte ou approximative
            if montant_str in tag_labels:
                scores['montant'] = self.weights['montant_match']
            elif any(montant_str[:3] in label for label in tag.get('labels', [])):
                scores['montant'] = self.weights['montant_match'] * 0.5
            else:
                scores['montant'] = 0
        
        # 3. Date/Année
        entry_years = set(entry.get('years', []))
        tag_years = set(tag.get('years', []))
        if entry_years and tag_years:
            year_overlap = len(entry_years & tag_years) / len(entry_years)
            scores['date'] = year_overlap * self.weights['date_match']
        else:
            scores['date'] = 0
        
        # 4. Nature
        if entry.get('nature_type') and tag.get('flux_type'):
            if (entry['nature_type'] == 'recette' and tag['flux_type'] == 'entree') or \
               (entry['nature_type'] == 'depense' and tag['flux_type'] == 'sortie'):
                scores['nature'] = self.weights['nature_match']
            else:
                scores['nature'] = 0
        else:
            scores['nature'] = 0
        
        # 5. Unité
        unite = str(entry.get('Unité', '')).lower()
        if unite and unite in tag_labels:
            scores['unite'] = self.weights['unite_match']
        else:
            scores['unite'] = 0
        
        # 6. Contexte source
        if 'SourcePhrase' in entry:
            source = str(entry['SourcePhrase']).lower()
            source_similarity = SequenceMatcher(None, source[:100], tag_labels[:100]).ratio()
            scores['source'] = source_similarity * self.weights['source_context']
        else:
            scores['source'] = 0
        
        # Score total
        total_score = sum(scores.values())
        
        return min(total_score, 1.0)
    
    async def _llm_select_best_tag_advanced(self, entry: Dict, 
                                          candidates: List[Dict]) -> Optional[Dict]:
        """Utilise le LLM avec toutes les informations pour sélectionner le meilleur tag"""
        if not candidates:
            return None
        
        # Préparer un prompt détaillé
        entry_info = f"""Entrée budgétaire complète:
- Axe: {entry.get('Axe', 'N/A')}
- Description: {entry.get('Description', 'N/A')}
- Montant: {entry.get('Montant', 'N/A')} {entry.get('Unité', '')}
- Date: {entry.get('Date', 'N/A')}
- Nature: {entry.get('Nature', 'N/A')}
- Contexte: {entry.get('SourcePhrase', 'N/A')[:100]}..."""
        
        tags_desc = []
        for i, candidate in enumerate(candidates):
            tag = candidate['tag']
            score = candidate['score']
            matches = ', '.join(candidate['matches'][:3])
            labels_preview = ', '.join(str(l)[:50] for l in tag.get('labels', [])[:3])
            
            tags_desc.append(
                f"{i}) {tag.get('sheet_name')}!{tag.get('cell_address')} "
                f"[Score: {score:.2f}, Matches: {matches}]\n   Labels: {labels_preview}"
            )
        
        system_prompt = """Tu es un expert en mapping budgétaire. Analyse toutes les informations fournies 
pour choisir la cellule la plus pertinente. Considère l'axe, la description, le montant, la date, 
la nature et le contexte. Réponds UNIQUEMENT avec le numéro (0-9) ou 'AUCUN'."""
        
        user_prompt = f"{entry_info}\n\nCellules candidates:\n" + "\n".join(tags_desc)
        
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ]
        
        try:
            response = await self.llm_client.chat(messages)
            if response:
                match = re.search(r'\b(\d)\b', response)
                if match:
                    idx = int(match.group(1))
                    if 0 <= idx < len(candidates):
                        selected_candidate = candidates[idx]
                        return self._create_detailed_mapping(
                            entry, 
                            selected_candidate['tag'],
                            selected_candidate['score'],
                            selected_candidate.get('matches', [])
                        )
        except Exception as e:
            logger.error(f"Erreur LLM: {str(e)}")
        
        # Fallback sur le meilleur score
        if candidates:
            best = candidates[0]
            return self._create_detailed_mapping(
                entry, best['tag'], best['score'], best.get('matches', [])
            )
        
        return None
    
    def _create_detailed_mapping(self, entry: Dict, tag: Dict, 
                               confidence_score: float, matches: List[str] = None) -> Dict:
        """Crée un mapping détaillé avec score de confiance"""
        return {
            "Axe": entry.get('Axe', ''),
            "Description": entry.get('Description', ''),
            "Montant": entry.get('Montant', 0),
            "Unite": entry.get('Unité', ''),
            "Date": entry.get('Date', ''),
            "Nature": entry.get('Nature', ''),
            "cellule": tag.get('cell_address', ''),
            "sheet_name": tag.get('sheet_name', 'Sheet1'),
            "tags_utilisés": tag.get('labels', [])[:3],
            "tag_id": tag.get('id', ''),
            "confidence_score": round(confidence_score, 3),
            "match_criteria": matches or [],
            "needs_review": confidence_score < 0.7
        }
    
    def _create_empty_mapping(self, entry: Dict) -> Dict:
        """Crée un mapping vide pour les entrées non mappées"""
        return {
            "Axe": entry.get('Axe', ''),
            "Description": entry.get('Description', ''),
            "Montant": entry.get('Montant', 0),
            "Unite": entry.get('Unité', ''),
            "Date": entry.get('Date', ''),
            "Nature": entry.get('Nature', ''),
            "cellule": "",
            "sheet_name": "",
            "tags_utilisés": [],
            "tag_id": "",
            "confidence_score": 0,
            "match_criteria": [],
            "needs_review": True,
            "not_mapped": True
        }
    
    def _get_cache_key(self, entry: Dict) -> str:
        """Génère une clé de cache unique pour une entrée"""
        return f"{entry.get('Axe', '')}|{entry.get('Description', '')}|{entry.get('Montant', '')}|{entry.get('Date', '')}"
    
    def _get_stop_words(self) -> Set[str]:
        """Retourne une liste étendue de mots vides"""
        return {
            # Articles et prépositions
            'dans', 'pour', 'avec', 'sans', 'sous', 'vers', 'chez',
            'entre', 'depuis', 'avant', 'après', 'pendant', 'contre',
            'selon', 'malgré', 'parmi', 'durant', 'concernant',
            # Mots communs
            'mois', 'année', 'jour', 'date', 'nombre', 'montant',
            'cellule', 'feuille', 'sheet', 'flux', 'effectifs',
            'cette', 'cela', 'celui', 'celle', 'ceux', 'celles',
            'tout', 'tous', 'toute', 'toutes', 'autre', 'autres',
            # Connecteurs
            'mais', 'donc', 'ainsi', 'alors', 'aussi', 'encore',
            'même', 'très', 'trop', 'plus', 'moins', 'bien'
        }
    
    def generate_mapping_report(self, mapping: List[Dict], 
                               entries_df: pd.DataFrame) -> Dict:
        """Génère un rapport détaillé du mapping pour vérification"""
        report = {
            'summary': {},
            'by_confidence': {},
            'by_sheet': {},
            'unmapped': [],
            'low_confidence': [],
            'details': []
        }
        
        # Statistiques générales
        total_entries = len(entries_df)
        mapped_entries = len([m for m in mapping if not m.get('not_mapped', False)])
        
        report['summary'] = {
            'total_entries': total_entries,
            'mapped_entries': mapped_entries,
            'unmapped_entries': total_entries - mapped_entries,
            'mapping_rate': (mapped_entries / total_entries * 100) if total_entries > 0 else 0,
            'average_confidence': sum(m.get('confidence_score', 0) for m in mapping) / len(mapping) if mapping else 0
        }
        
        # Répartition par niveau de confiance
        confidence_ranges = {
            'Très élevé (>90%)': 0,
            'Élevé (70-90%)': 0,
            'Moyen (50-70%)': 0,
            'Faible (<50%)': 0,
            'Non mappé': 0
        }
        
        for m in mapping:
            score = m.get('confidence_score', 0)
            if m.get('not_mapped', False):
                confidence_ranges['Non mappé'] += 1
            elif score > 0.9:
                confidence_ranges['Très élevé (>90%)'] += 1
            elif score > 0.7:
                confidence_ranges['Élevé (70-90%)'] += 1
            elif score > 0.5:
                confidence_ranges['Moyen (50-70%)'] += 1
            else:
                confidence_ranges['Faible (<50%)'] += 1
        
        report['by_confidence'] = confidence_ranges
        
        # Répartition par feuille
        sheet_counts = defaultdict(int)
        for m in mapping:
            if not m.get('not_mapped', False):
                sheet = m.get('sheet_name', 'Unknown')
                sheet_counts[sheet] += 1
        
        report['by_sheet'] = dict(sheet_counts)
        
        # Entrées non mappées
        for m in mapping:
            if m.get('not_mapped', False):
                report['unmapped'].append({
                    'description': m.get('Description', ''),
                    'montant': m.get('Montant', 0),
                    'axe': m.get('Axe', '')
                })
        
        # Entrées à faible confiance
        for m in mapping:
            if not m.get('not_mapped', False) and m.get('confidence_score', 0) < 0.7:
                report['low_confidence'].append({
                    'description': m.get('Description', ''),
                    'montant': m.get('Montant', 0),
                    'cellule': f"{m.get('sheet_name')}!{m.get('cellule')}",
                    'confidence': m.get('confidence_score', 0),
                    'matches': m.get('match_criteria', [])
                })
        
        # Détails complets pour export
        report['details'] = mapping
        
        return report
    
    def enrich_entries_with_mapping(self, entries: pd.DataFrame, 
                                   mapping: List[Dict]) -> pd.DataFrame:
        """Enrichit les entrées avec les informations de mapping détaillées"""
        # Créer un dictionnaire pour lookup rapide
        mapping_dict = {}
        for m in mapping:
            key = (m.get('Axe'), m.get('Description'), m.get('Montant'))
            mapping_dict[key] = m
        
        # Ajouter toutes les colonnes de mapping
        new_columns = [
            'CelluleCible', 'SheetName', 'TagID', 'TagsUtilises',
            'ConfidenceScore', 'NeedsReview', 'MatchCriteria', 'IsMapped'
        ]
        
        for col in new_columns:
            if col not in entries.columns:
                entries[col] = None
        
        # Enrichir
        for idx, row in entries.iterrows():
            key = (row.get('Axe'), row.get('Description'), row.get('Montant'))
            
            if key in mapping_dict:
                m = mapping_dict[key]
                entries.loc[idx, 'CelluleCible'] = m.get('cellule', '')
                entries.loc[idx, 'SheetName'] = m.get('sheet_name', '')
                entries.loc[idx, 'TagID'] = m.get('tag_id', '')
                entries.loc[idx, 'TagsUtilises'] = ', '.join(m.get('tags_utilisés', []))
                entries.loc[idx, 'ConfidenceScore'] = m.get('confidence_score', 0)
                entries.loc[idx, 'NeedsReview'] = m.get('needs_review', False)
                entries.loc[idx, 'MatchCriteria'] = ', '.join(m.get('match_criteria', []))
                entries.loc[idx, 'IsMapped'] = not m.get('not_mapped', False)
            else:
                entries.loc[idx, 'IsMapped'] = False
                entries.loc[idx, 'NeedsReview'] = True
        
        return entries