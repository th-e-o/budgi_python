# modules/bpss_tool.py - VERSION CORRIGÉE
import pandas as pd
import openpyxl
from typing import Optional, Dict, Any
import logging

logger = logging.getLogger(__name__)

class BPSSTool:
    """Outil BPSS pour traitement des fichiers budgétaires - VERSION CORRIGÉE"""
    
    def __init__(self):
        self.config = {
            'default_year': 2025,
            'default_ministry': '38',
            'default_program': '150'
        }
    
    def process_files(self, ppes_path: str, dpp18_path: str, bud45_path: str,
                 year: int, ministry_code: str, program_code: str,
                 target_workbook: openpyxl.Workbook) -> openpyxl.Workbook:
        """Traite les fichiers BPSS et met à jour le workbook cible"""
        try:
            # Vérifier que les fichiers existent
            import os
            for path, name in [(ppes_path, 'PP-E-S'), (dpp18_path, 'DPP18'), (bud45_path, 'BUD45')]:
                if not os.path.exists(path):
                    raise FileNotFoundError(f"Fichier {name} introuvable: {path}")
                logger.info(f"Fichier {name} trouvé: {path}")
            
            # Charger les données
            logger.info("Chargement des fichiers BPSS...")
            
            # PP-E-S - Gérer les erreurs de feuilles manquantes
            sheet_names = self._get_sheet_names(ministry_code, program_code)
            
            try:
                # Essayer de charger avec les noms de feuilles calculés
                df_pp_categ = pd.read_excel(ppes_path, sheet_name=sheet_names['pp_categ'])
                df_entrants = pd.read_excel(ppes_path, sheet_name=sheet_names['entrants'])
                df_sortants = pd.read_excel(ppes_path, sheet_name=sheet_names['sortants'])
            except Exception as e:
                logger.warning(f"Erreur avec les noms de feuilles calculés: {str(e)}")
                
                # Fallback : Lister les feuilles disponibles et essayer de deviner
                try:
                    xl_file = pd.ExcelFile(ppes_path)
                    available_sheets = xl_file.sheet_names
                    logger.info(f"Feuilles disponibles dans PP-E-S: {available_sheets}")
                    
                    # Chercher les feuilles par pattern
                    pp_categ_sheet = None
                    entrants_sheet = None
                    sortants_sheet = None
                    
                    for sheet in available_sheets:
                        sheet_lower = sheet.lower()
                        if 'pp_categ' in sheet_lower or 'categ' in sheet_lower:
                            pp_categ_sheet = sheet
                        elif 'entrant' in sheet_lower:
                            entrants_sheet = sheet
                        elif 'sortant' in sheet_lower:
                            sortants_sheet = sheet
                    
                    # Si on ne trouve pas, prendre les 3 premières feuilles
                    if not all([pp_categ_sheet, entrants_sheet, sortants_sheet]):
                        if len(available_sheets) >= 3:
                            pp_categ_sheet = pp_categ_sheet or available_sheets[0]
                            entrants_sheet = entrants_sheet or available_sheets[1]
                            sortants_sheet = sortants_sheet or available_sheets[2]
                        else:
                            raise ValueError(f"Le fichier PP-E-S doit contenir au moins 3 feuilles, trouvé: {len(available_sheets)}")
                    
                    logger.info(f"Utilisation des feuilles: {pp_categ_sheet}, {entrants_sheet}, {sortants_sheet}")
                    
                    # Charger avec les feuilles trouvées
                    df_pp_categ = pd.read_excel(ppes_path, sheet_name=pp_categ_sheet)
                    df_entrants = pd.read_excel(ppes_path, sheet_name=entrants_sheet)
                    df_sortants = pd.read_excel(ppes_path, sheet_name=sortants_sheet)
                    
                except Exception as e2:
                    logger.error(f"Impossible de charger PP-E-S: {str(e2)}")
                    raise
            
            # DPP18 et BUD45 - Plus robuste
            try:
                df_dpp18 = pd.read_excel(dpp18_path)
            except Exception as e:
                logger.error(f"Erreur chargement DPP18: {str(e)}")
                # Essayer de charger la première feuille
                df_dpp18 = pd.read_excel(dpp18_path, sheet_name=0)
            
            try:
                df_bud45 = pd.read_excel(bud45_path)
            except Exception as e:
                logger.error(f"Erreur chargement BUD45: {str(e)}")
                # Essayer de charger la première feuille
                df_bud45 = pd.read_excel(bud45_path, sheet_name=0)
                        
            # Appliquer les traitements
            target_workbook = self._completion_accueil(
                target_workbook, year, ministry_code, program_code
            )
            target_workbook = self._load_ppes_data(
                target_workbook, df_pp_categ, df_entrants, df_sortants, program_code
            )
            target_workbook = self._load_dpp18_data(
                target_workbook, df_dpp18, program_code
            )
            target_workbook = self._load_bud45_data(
                target_workbook, df_bud45, program_code
            )
            
            logger.info("Traitement BPSS terminé avec succès")
            return target_workbook
            
        except Exception as e:
            logger.error(f"Erreur traitement BPSS: {str(e)}")
            raise
    
    def _get_sheet_names(self, ministry_code: str, program_code: str) -> Dict[str, str]:
        """Génère les noms de feuilles selon les codes"""
        return {
            'pp_categ': f"MIN_{ministry_code}_DETAIL_Prog_PP_CATEG",
            'entrants': f"MIN_{ministry_code}_DETAIL_Prog_Entrants",
            'sortants': f"MIN_{ministry_code}_DETAIL_Prog_Sortants"
        }
    
    
    def _completion_accueil(self, wb: openpyxl.Workbook, year: int, ministry_code:str, program_code:str) -> openpyxl.Workbook:
        """Completion de la page accueil"""
        accueil_sheet = wb['Accueil']
            
        accueil_sheet.cell(row=34, column=3, value=year)
        accueil_sheet.cell(row=35, column=3, value=year+1)
            
        accueil_sheet.cell(row=37, column=4, value=ministry_code)
        accueil_sheet.cell(row=39, column=4, value=program_code)

        return wb


    def _load_ppes_data(self, wb: openpyxl.Workbook, df_pp_categ: pd.DataFrame,
                       df_entrants: pd.DataFrame, df_sortants: pd.DataFrame,
                       program_code: str) -> openpyxl.Workbook:
        """Charge les données PP-E-S dans le workbook - VERSION CORRIGÉE"""
        # Filtrer par programme (3 premiers caractères)
        code_prefix = program_code[:3]
        
        df1 = df_pp_categ[df_pp_categ['nom_prog'].astype(str).str[:3] == code_prefix]
        df2 = df_entrants[df_entrants['nom_prog'].astype(str).str[:3] == code_prefix]
        df3 = df_sortants[df_sortants['nom_prog'].astype(str).str[:3] == code_prefix]
        
        # Créer ou obtenir la feuille
        sheet_name = "Données PP-E-S"
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        
        sheet = wb[sheet_name]
        
        # Écrire les données - CORRECTION: utiliser colonne 2 (B) au lieu de 3 (C)
        start_rows = [7, 106, 205]  # Positions de départ
        dataframes = [df1, df2, df3]
        
        for start_row, df in zip(start_rows, dataframes):
            # Limiter à 100 lignes comme dans le VBA
            df_limited = df.head(100)
            
            for r_idx, row in enumerate(df_limited.values):
                for c_idx, value in enumerate(row):
                    sheet.cell(row=start_row + r_idx, column=3 + c_idx, value=value)
                    if start_row == 7 : #Pour la première feuille
                        sheet.cell(row = start_row + r_idx, column = 2, value = row[3][:4])
                    else : #Pour les deux suivantes
                        sheet.cell(row = start_row + r_idx, column = 2, value = row[2][:4])

        # Traitement spécial "Indicié" 
        
        # Traitement spécial "Indicié" - CORRECTION v2
        # La colonne "marqueur_masse_indiciaire" pourrait être à différents endroits
        # Essayons plusieurs colonnes possibles
        df_indicie = None
        
        # Chercher la colonne qui contient "Indicié"
        for col in df1.columns:
            if 'Indicié' in df1[col].astype(str).values:
                df_indicie = df1[df1[col] == 'Indicié']
                logger.info(f"Colonne 'Indicié' trouvée : {col}")
                break
        
        if df_indicie is None or df_indicie.empty:
            logger.warning("Aucune donnée 'Indicié' trouvée dans PP-E-S")
        else:
            if 'Accueil' not in wb.sheetnames:
                wb.create_sheet('Accueil')
            
            accueil_sheet = wb['Accueil']
            
            # Nettoyer d'abord les anciennes données
            for row in range(43, 55):  # B43:C54
                accueil_sheet.cell(row=row, column=2, value=None)  # Colonne B
                accueil_sheet.cell(row=row, column=3, value=None)  # Colonne C
            
            # Nettoyer d'abord les anciennes données 
            for row in range(43,54):
                accueil_sheet.cell(row=row, column=2, value=None)
                accueil_sheet.cell(row=row, column=3, value=None)

            # Écrire les nouvelles données
            row_dest = 43
            for idx, row_data in enumerate(df_indicie.values):
                if row_dest > 54:  # Ne pas dépasser la ligne 54
                    break
                
                code_categorie = row_data[3][:4]
                nom_categorie = row_data[3]
                
                # Ecrire dans Accueil si on a trouvé les données
                if code_categorie and nom_categorie:
                    accueil_sheet.cell(row=row_dest, column=2,value=code_categorie)
                    accueil_sheet.cell(row=row_dest, column=3, value=nom_categorie)
                    row_dest += 1
                else:
                    logger.warning(f"Impossible d'extraire code/nom de la ligne: {row_data[:5]}")

        logger.info("Données PP-E-S chargées")
        return wb
    
    def _load_dpp18_data(self, wb: openpyxl.Workbook, df_dpp18: pd.DataFrame,
                        program_code: str) -> openpyxl.Workbook:
        """Charge les données DPP18 dans le workbook - VERSION CORRIGÉE"""
        sheet_name = "INF DPP 18"
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        
        sheet = wb[sheet_name]
        
        # Header (5 premières lignes) - écrire à partir de colonne B
        header = df_dpp18.head(5)
        for r_idx, row in enumerate(header.values):
            for c_idx, value in enumerate(row):
                sheet.cell(row=2 + r_idx, column=1 + c_idx, value=value)
        
        # Filtrage par programme sur la colonne A (index 0)
        code_prefix = program_code[:3]
        df_filtered = df_dpp18[
            df_dpp18.iloc[:, 0].astype(str).str.contains(code_prefix, na=False)
        ]
        
        # Écrire les données filtrées à partir de la ligne 6
        for r_idx, row in enumerate(df_filtered.values):
            for c_idx, value in enumerate(row):
                sheet.cell(row = 6 + r_idx, column = 1, value = row[1][:14])
        
        logger.info("Données DPP18 chargées")
        return wb
    
    def _load_bud45_data(self, wb: openpyxl.Workbook, df_bud45: pd.DataFrame,
                        program_code: str) -> openpyxl.Workbook:
        """Charge les données BUD45 dans le workbook - VERSION CORRIGÉE"""
        sheet_name = "INF BUD 45"
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        
        sheet = wb[sheet_name]
        
        # Header (5 premières lignes) - écrire à partir de colonne B
        header = df_bud45.head(5)
        for r_idx, row in enumerate(header.values):
            for c_idx, value in enumerate(row):
                sheet.cell(row = 2 + r_idx, column = 1 + c_idx, value=value)
        
        # Filtrage par programme sur la colonne B (index 1)
        code_prefix = program_code[:3]
        if len(df_bud45.columns) > 1:
            df_filtered = df_bud45[
                df_bud45.iloc[:, 1].astype(str).str.contains(code_prefix, na=False)
            ]
            
            # Écrire les données filtrées à partir de la ligne 6
            for r_idx, row in enumerate(df_filtered.values):
                for c_idx, value in enumerate(row):
                    sheet.cell(row = 7 + r_idx, column = 1, value = row[3][:14])
        
        logger.info("Données BUD45 chargées")
        return wb