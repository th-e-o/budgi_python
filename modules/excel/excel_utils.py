from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles.proxy import StyleProxy
from openpyxl.worksheet.worksheet import Worksheet
from colorsys import rgb_to_hls, hls_to_rgb

RGBMAX = 0xff
HLSMAX = 240

class ExcelUtils:
    @staticmethod
    def get_data_only_range(sheet: Worksheet) -> tuple[int, int]:
        """
        Get the range of cells that actually contain values (not just formatting).
        This is more restrictive than max_row/max_column.
        """
        last_row_with_data = 0
        last_col_with_data = 0

        max_row = sheet.max_row
        max_col = sheet.max_column

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row, col)
                if cell.value is not None:
                    last_row_with_data = row
                    if col > last_col_with_data:
                        last_col_with_data = col

        return last_row_with_data, last_col_with_data

    @staticmethod
    def _rgb_to_ms_hls(red, green=None, blue=None):
        """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
        if green is None:
            if isinstance(red, str):
                if len(red) > 6:
                    red = red[-6:]  # Ignore preceding '#' and alpha values
                blue = int(red[4:], 16) / RGBMAX
                green = int(red[2:4], 16) / RGBMAX
                red = int(red[0:2], 16) / RGBMAX
            else:
                red, green, blue = red
        h, l, s = rgb_to_hls(red, green, blue)
        return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))

    @staticmethod
    def _ms_hls_to_rgb(hue, lightness=None, saturation=None):
        """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
        if lightness is None:
            hue, lightness, saturation = hue
        return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)

    @staticmethod
    def _rgb_to_hex(red, green=None, blue=None):
        """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
        if green is None:
            red, green, blue = red
        return ('%02x%02x%02x' % (
        int(round(red * RGBMAX)), int(round(green * RGBMAX)), int(round(blue * RGBMAX)))).upper()

    @staticmethod
    def get_theme_colors(wb):
        """Gets theme colors from the workbook"""
        from openpyxl.xml.functions import QName, fromstring
        xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
        root = fromstring(wb.loaded_theme)
        themeEl = root.find(QName(xlmns, 'themeElements').text)
        colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
        firstColorScheme = colorSchemes[0]

        colors = []

        for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
            accent = firstColorScheme.find(QName(xlmns, c).text)

            if 'window' in accent.getchildren()[0].attrib['val']:
                colors.append(accent.getchildren()[0].attrib['lastClr'])
            else:
                colors.append(accent.getchildren()[0].attrib['val'])

        return colors

    @staticmethod
    def _tint_luminance(tint, lum):
        """Tints a HLSMAX based luminance"""
        # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
        if tint < 0:
            return int(round(lum * (1.0 + tint)))
        else:
            return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))

    @staticmethod
    def theme_and_tint_to_rgb(theme_colors: List[str], theme: int, tint: float) -> str:
        rgb = theme_colors[theme]
        h, l, s = ExcelUtils._rgb_to_ms_hls(rgb)
        return ExcelUtils._rgb_to_hex(ExcelUtils._ms_hls_to_rgb(h, ExcelUtils._tint_luminance(tint, l), s))

    @staticmethod
    def extract_cell_colors(fill, font, theme_colors: List[str]) -> Tuple[str, str]:
        """
        Extracts the background and font color from a cell
        """
        base_fill_color = 'FFFFFF'  # Default to white if no fill color is found
        if isinstance(fill, StyleProxy) and (fill.patternType is None or fill.patternType == "solid"):
            # FgColor is the solid fill color. Apply its tint.
            fgColor = fill.fgColor
            bgColor = fill.bgColor
            if fgColor:
                if fgColor.type == 'rgb':
                    if fgColor.rgb == '00000000':  # Special case for transparent fill'
                        base_fill_color = 'FFFFFF'  # Default to white for transparent fill
                    else:
                        base_fill_color = fgColor.rgb[-6:]
                elif fgColor.type == 'theme':
                    base_fill_color = ExcelUtils.theme_and_tint_to_rgb(theme_colors, fgColor.theme, fgColor.tint)
            elif bgColor:
                if bgColor.type == 'rgb':
                    base_fill_color = bgColor.rgb[-6:]
                elif bgColor.type == 'theme':
                    base_fill_color = ExcelUtils.theme_and_tint_to_rgb(theme_colors, bgColor.theme, bgColor.tint)
            else:
                base_fill_color = 'FFFFFF'

        font_color_hex = '000000'  # Default to black if no color is set
        if font.color is not None:
            if font.color.type == 'rgb':
                font_color_hex = font.color.rgb[-6:]
            elif font.color.type == 'theme':
                font_color_hex = ExcelUtils.theme_and_tint_to_rgb(theme_colors, font.color.theme, font.color.tint)

        return base_fill_color, font_color_hex


    @staticmethod
    def extract_border_color(side, theme_colors):
        """Extract color from a border side, handling theme colors and RGB"""
        if side is None or side.color is None:
            return None

        color = side.color
        if color.type == 'rgb' and color.rgb:
            return color.rgb[-6:]  # Remove FF prefix if present
        elif color.type == 'theme' and color.theme is not None:
            return ExcelUtils.theme_and_tint_to_rgb(theme_colors, color.theme, color.tint or 0)
        elif color.type == 'indexed':
            # Handle indexed colors if needed - would require color palette lookup
            return None
        else:
            return None

    @staticmethod
    def extract_cell_borders(cell, theme_colors):
        """
        Extract border information from a cell.
        Returns a dictionary with border information for each side.
        """
        if not hasattr(cell, 'border') or cell.border is None:
            return None

        border = cell.border
        border_info = {}

        # Define the sides to check
        sides = {
            'top': border.top,
            'bottom': border.bottom,
            'left': border.left,
            'right': border.right,
            'diagonal': border.diagonal
        }

        # Extract information for each side
        for side_name, side in sides.items():
            if side is not None and side.style is not None:
                side_info = {
                    'style': side.style,
                    'color': ExcelUtils.extract_border_color(side, theme_colors)
                }
                border_info[side_name] = side_info

        # Add diagonal direction info if diagonal border exists
        if border.diagonal is not None and border.diagonal.style is not None:
            border_info['diagonal_up'] = border.diagonalUp
            border_info['diagonal_down'] = border.diagonalDown

        return border_info if border_info else None

    @staticmethod
    def get_column_letter(column_index: int) -> str:
        """
        Converts a 1-based column index into an Excel column letter.
        (e.g., 1 -> 'A', 27 -> 'AA').
        """
        if not isinstance(column_index, int) or column_index < 1:
            raise ValueError("Column index must be a positive integer.")

        string = ""
        while column_index > 0:
            column_index, remainder = divmod(column_index - 1, 26)
            string = chr(65 + remainder) + string
        return string