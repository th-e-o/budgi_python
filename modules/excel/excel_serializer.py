from __future__ import annotations

import dataclasses
import json
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, Any, List, Tuple, Set

import json5
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from modules.excel.excel_utils import ExcelUtils


@dataclass
class FontData:
    font: str = None
    font_size: float = None
    bold: bool = False
    italic: bool = False
    underline: bool = False
    strike: bool = False
    color: str = None


@dataclass
class BorderData:
    style: str = None
    color: str = None


@dataclass
class CellStyle:
    background_color: str = None
    borders: Dict[str, BorderData] = None


@dataclass
class CellData:
    value: any
    cell_type: str
    font: FontData = None
    align_horizontal: str = "left"
    align_vertical: str = None
    wrap_text: bool = False
    style: CellStyle = None


@dataclass
class SheetData:
    name: str
    cells: Dict[str, CellData] = None
    dimensions: tuple[int, int] = (0, 0)
    row_heights: Dict[int, float] = None
    column_widths: Dict[str, float] = None
    merged_ranges: list[str] = None


class ExcelSerializer:
    def __init__(self, workbook: Workbook = None):
        """Initialize the ExcelSerializer with an optional openpyxl Workbook."""
        self.workbook = workbook if workbook else Workbook()
        self.sheet_real_dimensions: Dict[str, tuple[int, int]] = self._initialize_sheet_dimensions()

    @staticmethod
    def from_workbook(workbook: Workbook) -> ExcelSerializer:
        """Convert an openpyxl Workbook to an ExcelSerializer instance."""
        serializer = ExcelSerializer(workbook)
        return serializer

    def _coord_to_rowcol(self, coord: str) -> Tuple[int, int]:
        """Convert Excel coordinate (e.g., 'A1') to (row, col) tuple (1-based)."""
        from openpyxl.utils.cell import coordinate_from_string
        from openpyxl.utils import column_index_from_string
        col, row = coordinate_from_string(coord)
        return row, column_index_from_string(col)

    def _is_cell_in_merged_range_but_not_topleft(self, row: int, col: int, merged_ranges) -> bool:
        """Check if cell is part of merged range but not the top-left cell."""
        for mr in merged_ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                if (row, col) != (mr.min_row, mr.min_col):
                    return True
        return False

    def _extract_cell_properties(self, cell, theme_colors) -> Dict[str, Any]:
        """Extract all properties from a cell."""
        background_color, font_color = ExcelUtils.extract_cell_colors(cell, theme_colors)
        borders = ExcelUtils.extract_cell_borders(cell, theme_colors)

        # Convert borders to our format
        border_data = None
        if borders:
            border_data = {}
            for side in ['top', 'bottom', 'left', 'right', 'diagonal']:
                if side in borders:
                    border_data[side] = {
                        'style': borders[side]['style'],
                        'color': borders[side]['color']
                    }

        # Handle cell value
        cell_value = cell.value
        if isinstance(cell_value, datetime):
            cell_value = cell_value.isoformat()
        elif not isinstance(cell_value, (str, int, float, bool, type(None))):
            cell_value = str(cell_value) if hasattr(cell_value, '__str__') else None

        return {
            'value': cell_value,
            'cell_type': cell.data_type,
            'bg_color': background_color,
            'font_name': cell.font.name,
            'font_size': cell.font.sz,
            'font_bold': cell.font.bold or False,
            'font_italic': cell.font.italic or False,
            'font_underline': cell.font.underline or False,
            'font_strike': cell.font.strike or False,
            'font_color': font_color,
            'align_horizontal': cell.alignment.horizontal or "left",
            'align_vertical': cell.alignment.vertical,
            'wrap_text': cell.alignment.wrap_text or False,
            'borders': border_data
        }

    def _optimize_ranges_for_attribute(self, cell_value_map: Dict[str, Any]) -> Dict[str, Any]:
        """Optimize ranges for a specific attribute."""
        if not cell_value_map:
            return {}

        # Group cells by value (handle both simple values and complex objects)
        value_groups: Dict[Any, List[Tuple[int, int]]] = {}
        for coord, value in cell_value_map.items():
            row, col = self._coord_to_rowcol(coord)
            # Convert complex objects to hashable representation for grouping
            if isinstance(value, dict):
                hashable_value = tuple(sorted(value.items()))
            else:
                hashable_value = value

            if hashable_value not in value_groups:
                value_groups[hashable_value] = []
            value_groups[hashable_value].append((row, col))

        # Find optimal ranges for each value group
        result = {}
        for hashable_value, cells in value_groups.items():
            ranges = self._find_optimal_ranges(cells)
            # Convert back to original value format
            if isinstance(hashable_value, tuple) and len(hashable_value) > 0 and isinstance(hashable_value[0], tuple):
                # This was a dict that we converted to tuple of tuples
                original_value = dict(hashable_value)
            else:
                original_value = hashable_value

            for range_str in ranges:
                result[range_str] = original_value

        return result

    def _find_optimal_ranges(self, cells: List[Tuple[int, int]]) -> List[str]:
        """Find optimal rectangular ranges that cover the given cells."""
        if not cells:
            return []

        cell_set = set(cells)
        used = set()
        ranges = []

        # Sort cells by row, then column for consistent processing
        sorted_cells = sorted(cells)

        for start_row, start_col in sorted_cells:
            if (start_row, start_col) in used:
                continue

            # Find the largest rectangle starting from this cell
            max_row, max_col = self._find_largest_rectangle(
                start_row, start_col, cell_set, used
            )

            # Mark all cells in this rectangle as used
            for r in range(start_row, max_row + 1):
                for c in range(start_col, max_col + 1):
                    used.add((r, c))

            # Convert to Excel range format
            range_str = self._format_range(start_row, start_col, max_row, max_col)
            ranges.append(range_str)

        return ranges

    def _find_largest_rectangle(self, start_row: int, start_col: int,
                                cell_set: Set[Tuple[int, int]], used: Set[Tuple[int, int]]) -> Tuple[int, int]:
        """Find the largest rectangle starting from the given position."""
        max_row = start_row
        max_col = start_col

        # First, expand right as much as possible in the starting row
        while (start_row, max_col + 1) in cell_set and (start_row, max_col + 1) not in used:
            max_col += 1

        # Then try to expand down while maintaining the rectangle width
        can_expand_down = True
        while can_expand_down:
            # Check if all cells in the next row (within our width) are available
            for c in range(start_col, max_col + 1):
                if (max_row + 1, c) not in cell_set or (max_row + 1, c) in used:
                    can_expand_down = False
                    break
            if can_expand_down:
                max_row += 1

        return max_row, max_col

    def _format_range(self, start_row: int, start_col: int, end_row: int, end_col: int) -> str:
        """Format a range into Excel notation."""
        start_coord = f"{ExcelUtils.get_column_letter(start_col)}{start_row}"
        if start_row == end_row and start_col == end_col:
            return start_coord
        else:
            end_coord = f"{ExcelUtils.get_column_letter(end_col)}{end_row}"
            return f"{start_coord}:{end_coord}"

    def serialize_to_data_optimized(self) -> Dict[str, Any]:
        """Serialize workbook to optimized attribute-based format."""
        theme_colors = ExcelUtils.get_theme_colors(self.workbook)
        result = {}

        for sheet in self.workbook.worksheets:
            sheet_name = sheet.title
            max_row, max_col = self.sheet_real_dimensions[sheet_name]

            # Get merged cell ranges
            merged_ranges = list(sheet.merged_cells.ranges)

            # Extract properties from all relevant cells
            cell_properties = {}
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    # Skip cells that are part of merged ranges but not top-left
                    if self._is_cell_in_merged_range_but_not_topleft(r, c, merged_ranges):
                        continue

                    cell = sheet.cell(row=r, column=c)
                    coord = f"{ExcelUtils.get_column_letter(c)}{r}"
                    cell_properties[coord] = self._extract_cell_properties(cell, theme_colors)

            # Create attribute maps
            attribute_maps = {}

            # Values with types (skip None/empty values)
            values_with_types_map = {}
            for coord, props in cell_properties.items():
                if props['value'] is not None and props['value'] != '':
                    values_with_types_map[coord] = {
                        'value': props['value'],
                        'type': props['cell_type']
                    }
            attribute_maps['valuesWithTypes'] = self._optimize_ranges_for_attribute(values_with_types_map)

            # Background colors (skip white/default)
            bg_color_map = {}
            for coord, props in cell_properties.items():
                if props['bg_color'] and props['bg_color'] != 'FFFFFF':
                    bg_color_map[coord] = f"#{props['bg_color']}"
            attribute_maps['bgColor'] = self._optimize_ranges_for_attribute(bg_color_map)

            # Font properties
            font_bold_map = {}
            font_italic_map = {}
            font_size_map = {}
            font_color_map = {}
            font_name_map = {}
            for coord, props in cell_properties.items():
                if props['font_bold']:
                    font_bold_map[coord] = True
                if props['font_italic']:
                    font_italic_map[coord] = True
                if props['font_size'] and props['font_size'] != 11:  # Skip default size
                    font_size_map[coord] = props['font_size']
                if props['font_color'] and props['font_color'] != '000000':  # Skip black
                    font_color_map[coord] = f"#{props['font_color']}"
                if props['font_name'] and props['font_name'] != 'Calibri':  # Skip default font
                    font_name_map[coord] = props['font_name']

            attribute_maps['fontBold'] = self._optimize_ranges_for_attribute(font_bold_map)
            attribute_maps['fontItalic'] = self._optimize_ranges_for_attribute(font_italic_map)
            attribute_maps['fontSize'] = self._optimize_ranges_for_attribute(font_size_map)
            attribute_maps['fontColor'] = self._optimize_ranges_for_attribute(font_color_map)
            attribute_maps['fontName'] = self._optimize_ranges_for_attribute(font_name_map)

            # Alignment properties
            align_h_map = {}
            align_v_map = {}
            wrap_text_map = {}
            for coord, props in cell_properties.items():
                if props['align_horizontal'] and props['align_horizontal'] != 'left':
                    align_h_map[coord] = props['align_horizontal']
                if props['align_vertical']:
                    align_v_map[coord] = props['align_vertical']
                if props['wrap_text']:
                    wrap_text_map[coord] = True

            attribute_maps['alignHorizontal'] = self._optimize_ranges_for_attribute(align_h_map)
            attribute_maps['alignVertical'] = self._optimize_ranges_for_attribute(align_v_map)
            attribute_maps['wrapText'] = self._optimize_ranges_for_attribute(wrap_text_map)

            # Borders (keep as individual cells - no range optimization)
            borders_map = {}
            for coord, props in cell_properties.items():
                if props['borders']:
                    borders_map[coord] = props['borders']
            attribute_maps['borders'] = borders_map

            # Compile sheet data
            result[sheet_name] = {
                **attribute_maps,
                'dimensions': self.sheet_real_dimensions[sheet_name],
                'rowHeights': self._extract_row_heights(sheet_name),
                'columnWidths': self._extract_column_widths(sheet_name),
                'mergedRanges': [str(mr) for mr in merged_ranges],
                'hiddenGridLines': sheet.sheet_view.showGridLines is False,
            }

        return result

    def serialize_to_json(self) -> Dict:
        """Serialize the workbook to a JSON dict with optimized structure."""
        return self.serialize_to_data_optimized()

    # Keep existing helper methods
    def _initialize_sheet_dimensions(self) -> Dict[str, tuple[int, int]]:
        """Initialize the dimensions of all sheets in the workbook."""
        sheet_dimensions = {}
        for sheet in self.workbook.worksheets:
            sheet_dimensions[sheet.title] = ExcelUtils.get_data_only_range(sheet)
        return sheet_dimensions

    def _extract_row_heights(self, sheet_name: str) -> Dict[int, float]:
        """Extract row heights from the sheet."""
        sheet: Worksheet = self.workbook[sheet_name]
        dimensions: Dict[int, float] = {}
        max_rows = self.sheet_real_dimensions[sheet_name][0]
        for row, dimension in sheet.row_dimensions.items():
            if row > max_rows:
                break
            if dimension.customHeight:
                dimensions[row] = dimension.ht
        return dimensions

    def _extract_column_widths(self, sheet_name: str) -> Dict[str, float]:
        """Extract column widths from the sheet."""
        sheet: Worksheet = self.workbook[sheet_name]
        dimensions: Dict[str, float] = {}
        max_columns = self.sheet_real_dimensions[sheet_name][1]
        for i, (col, dimension) in enumerate(sheet.column_dimensions.items()):
            if i > max_columns:
                break
            if dimension.width:
                dimensions[col] = dimension.width
        return dimensions

    # Keep the original methods for backward compatibility
    def serialize_to_data(self) -> Dict[str, SheetData]:
        """Original serialize method for backward compatibility."""
        theme_colors = ExcelUtils.get_theme_colors(self.workbook)
        sheet_datas = {}
        for sheet in self.workbook.worksheets:
            sheet_name = sheet.title
            sheet_datas[sheet_name] = SheetData(name=sheet_name, cells={})

            row_heights = self._extract_row_heights(sheet_name)
            column_widths = self._extract_column_widths(sheet_name)

            for row_number, row in enumerate(sheet.iter_rows()):
                for cell_number, cell in enumerate(row):
                    if row_number >= self.sheet_real_dimensions[sheet_name][0] or \
                            cell_number >= self.sheet_real_dimensions[sheet_name][1]:
                        break

                    cell_coordinates = cell.coordinate
                    background_color, font_color = ExcelUtils.extract_cell_colors(cell, theme_colors)
                    borders = ExcelUtils.extract_cell_borders(cell, theme_colors)
                    borderdatas = {
                        'top': BorderData(),
                        'bottom': BorderData(),
                        'left': BorderData(),
                        'right': BorderData(),
                        'diagonal': BorderData()
                    }

                    if borders:
                        for side, b in borders.items():
                            if side in borderdatas:
                                borderdatas[side].style = b['style']
                                borderdatas[side].color = b['color']

                    cell_style = CellStyle(borders=borderdatas, background_color=background_color)
                    font_data = FontData(
                        font=cell.font.name,
                        font_size=cell.font.sz,
                        bold=cell.font.bold or False,
                        italic=cell.font.italic or False,
                        underline=cell.font.underline or False,
                        strike=cell.font.strike or False,
                        color=font_color
                    )

                    cell_value = cell.value
                    if isinstance(cell_value, datetime):
                        cell_value = cell_value.isoformat()

                    if not isinstance(cell_value, (str, int, float, bool, type(None))):
                        cell_value = cell_value.text if hasattr(cell_value, 'text') else str(cell_value)

                    cell_data = CellData(
                        value=cell_value,
                        cell_type=cell.data_type,
                        font=font_data,
                        align_horizontal=cell.alignment.horizontal or "left",
                        align_vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text,
                        style=cell_style
                    )

                    if sheet_name not in sheet_datas:
                        sheet_datas[sheet_name] = SheetData(name=sheet_name, cells={})

                    sheet_datas[sheet_name].cells[cell_coordinates] = cell_data

            sheet_datas[sheet_name].dimensions = self.sheet_real_dimensions[sheet_name]
            sheet_datas[sheet_name].row_heights = row_heights
            sheet_datas[sheet_name].column_widths = column_widths
            sheet_datas[sheet_name].merged_ranges = [str(mr) for mr in sheet.merged_cells.ranges]

        return sheet_datas