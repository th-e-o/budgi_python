import logging
from collections import defaultdict
from enum import Enum
from io import BytesIO
from typing import List, Dict, Set, Optional

import openpyxl
from openpyxl.workbook import Workbook

from backend.core.excel_handler.excel_handler import UpdatedExcelHandler
from backend.core.excel_handler.operation_types import BackendOperationType, FrontendOperationType
from core.ExcelToUniverConverterOpt import ExcelToUniverConverterOpt

logger = logging.getLogger(__name__)


class ExcelUpdateCompiler:
    """
    Analyzes a list of Excel operations and compiles numerous cell updates
    for a single sheet into a more efficient, single 'REPLACE_SHEET' operation.
    """

    def __init__(self, handler: UpdatedExcelHandler, cell_update_threshold: int = 50):
        """
        Initializes the compiler.

        Args:
            handler: The Excel handler that holds the current state of the workbook.
            cell_update_threshold: The number of cell updates on a single sheet
                                   to trigger compilation for that sheet.
        """
        self.handler = handler
        self.threshold = cell_update_threshold
        self.temp_workbook: Optional[Workbook] = None

    def compile(self, operations: List[Dict]) -> List[Dict]:
        """
        Takes a list of operations and returns a new, potentially smaller list
        where heavy modifications have been compiled. This version modifies sheets
        directly within an in-memory copy of the workbook.

        Args:
            operations: The original list of operation dictionaries.

        Returns:
            A new list of optimized operations.
        """
        if not self.handler.has_workbook():
            logger.warning("Compiler cannot run without a base workbook. Skipping.")
            return operations

        # 1. Analyze which sheets have enough cell updates to be compiled
        sheets_to_compile = self._find_sheets_to_compile(operations)
        if not sheets_to_compile:
            return operations  # No optimization needed

        logger.info(f"Compiler will optimize updates for sheets: {list(sheets_to_compile)}")

        # 2. Create a full, independent copy of the workbook to modify safely
        workbook_bytes = self.handler.save_workbook_to_bytes(self.handler.workbook)
        with BytesIO(workbook_bytes) as stream:
            self.temp_workbook = openpyxl.load_workbook(BytesIO(workbook_bytes))

        # 3. Process operations: apply changes to sheets in the temp workbook or pass them through
        final_ops: List[Dict] = []
        for op in operations:
            op_sheet = op['handler_payload'].get('sheet_name')
            if op_sheet and op_sheet in sheets_to_compile:
                # If this is a cell update for a compiled sheet, apply it to the temp workbook.
                # All other operations on this sheet (like delete, create) are effectively
                # consumed by the final REPLACE_SHEET operation, so we skip them.
                if op['backend_type'] == BackendOperationType.UPDATE_CELL_VALUE:
                    self._apply_update_to_temp_sheet(op)
            else:
                # This operation is not on a sheet being compiled, so pass it through.
                final_ops.append(op)

        # 4. Generate the new 'REPLACE_SHEET' operations from the modified temp workbook
        compiled_ops = self._generate_compiled_ops(sheets_to_compile)

        def enum_to_json(obj):
            if isinstance(obj, Enum):
                return obj.value
            if isinstance(obj, bytes):
                return "<bytes>"
            raise TypeError(f"Object of type {type(obj)} is not JSON serializable")

        with open('compiled_operations.json', 'w') as f:
            import json
            json.dump(compiled_ops, f, default=enum_to_json, indent=4)

        final_ops.extend(compiled_ops)

        self.temp_workbook = None  # Release the temporary workbook from memory

        logger.info(f"Compilation complete. Original ops: {len(operations)}, Compiled ops: {len(final_ops)}")
        return final_ops

    def _find_sheets_to_compile(self, operations: List[Dict]) -> Set[str]:
        """Counts cell updates per sheet and identifies which exceed the threshold."""
        update_counts = defaultdict(int)
        for op in operations:
            if op['backend_type'] == BackendOperationType.UPDATE_CELL_VALUE:
                sheet_name = op['handler_payload'].get('sheet_name')
                if sheet_name:
                    update_counts[sheet_name] += 1
        return {sheet for sheet, count in update_counts.items() if count > self.threshold}

    def _apply_update_to_temp_sheet(self, op: Dict):
        """Applies a single UPDATE_CELL operation directly to a sheet in the temp workbook."""
        payload = op['handler_payload']
        sheet_name = payload['sheet_name']

        if sheet_name not in self.temp_workbook.sheetnames:
            # This case handles when a sheet is created and then heavily modified
            # within the same transaction. We must create it in our temp workbook too.
            self.temp_workbook.create_sheet(sheet_name)

        temp_sheet = self.temp_workbook[sheet_name]
        # Payload uses 0-based index, openpyxl uses 1-based
        row, col = payload['row'] + 1, payload['column'] + 1
        temp_sheet.cell(row=row, column=col, value=payload['value'])

    def _generate_compiled_ops(self, sheets_to_compile: Set[str]) -> List[Dict]:
        """
        Creates the final REPLACE_SHEET operations from the modified sheets
        in the temporary workbook.
        """
        # Save the entire modified temp workbook to bytes once for efficiency
        with BytesIO() as bio:
            self.temp_workbook.save(bio)
            source_workbook_bytes = bio.getvalue()

        converter = ExcelToUniverConverterOpt(self.temp_workbook)
        compiled_ops = []

        for sheet_name in sheets_to_compile:
            temp_sheet = self.temp_workbook[sheet_name]

            # Generate the UI payload (Univer format) for the modified sheet
            ui_sheet_data = converter.convert_sheet(temp_sheet, explicit_styles=True)
            ui_sheet_data['name'] = sheet_name  # Ensure the UI replaces the correct sheet

            # Build the new compiled operation
            op = {
                "id": f"compiled-op-{sheet_name}",
                "frontend_type": FrontendOperationType.REPLACE_SHEET,
                "backend_type": BackendOperationType.REPLACE_SHEET_FROM_ANOTHER_WORKBOOK,
                "description": f"Mettre à jour en bloc la feuille '{sheet_name}'",
                "handler_payload": {
                    "source_workbook_bytes": source_workbook_bytes,
                    "sheet_name": sheet_name,
                    "new_sheet_name": sheet_name
                },
                "ui_payload": ui_sheet_data
            }
            compiled_ops.append(op)

        return compiled_ops