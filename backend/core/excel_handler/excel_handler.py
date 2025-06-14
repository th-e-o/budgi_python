import copy

import openpyxl
import pandas as pd
from io import BytesIO
from typing import Optional, Any, Dict, List
import logging

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from backend.core.excel_handler.operation_types import BackendOperationType


class UpdatedExcelHandler:
    """
    Manages the in-memory state of the Excel workbook.
    Maintains an original and a user-modified version for non-destructive workflows.
    Acts as the execution engine for backend operations.
    """
    logger = logging.getLogger("UpdatedExcelHandler")

    def __init__(self):
        self._original_workbook: Optional[openpyxl.Workbook] = None
        self._user_workbook: Optional[openpyxl.Workbook] = None

    @property
    def workbook(self) -> Optional[openpyxl.Workbook]:
        """Returns the current, user-modified workbook."""
        return self._user_workbook

    def has_workbook(self) -> bool:
        return self._original_workbook is not None

    def load_workbook_from_bytes(self, file_bytes: bytes):
        """Loads a workbook, creating both the original and user-modifiable copies."""
        with BytesIO(file_bytes) as stream:
            self._original_workbook = openpyxl.load_workbook(stream, data_only=True, keep_vba=False)
            stream.seek(0)
            self._user_workbook = openpyxl.load_workbook(stream, data_only=True, keep_vba=False)
        self.logger.info("New workbook loaded. Original and user copies created.")

    def reset_to_original(self):
        """Discards all user changes and restores the workbook to its original state."""
        if not self._original_workbook:
            self.logger.warning("Cannot reset, no original workbook loaded.")
            return

        workbook_as_bytes = self.save_workbook_to_bytes(self._original_workbook)
        self._original_workbook = openpyxl.load_workbook(BytesIO(workbook_as_bytes), data_only=True, keep_vba=False)
        self.logger.info("Workbook has been reset to its original state.")

    def apply_updates(self, operations: List[Dict[str, Any]]):
        """Applies a list of backend operations to the user workbook."""
        if not self.has_workbook():
            raise ValueError("No workbook loaded to apply updates.")

        for op in operations:
            op_type = op.get('type')
            payload = op.get('payload', {})
            self.logger.info(f"Applying operation: {op_type.name} - {op.get('description')}")

            try:
                if op_type == BackendOperationType.CREATE_SHEET:
                    if payload['sheet_name'] not in self._user_workbook.sheetnames:
                        self._user_workbook.create_sheet(title=payload['sheet_name'])

                elif op_type == BackendOperationType.DELETE_SHEET:
                    if payload['sheet_name'] in self._user_workbook.sheetnames:
                        del self._user_workbook[payload['sheet_name']]

                elif op_type == BackendOperationType.UPDATE_CELL_VALUE:
                    sheet = self._user_workbook[payload['sheet_name']]
                    cell = sheet.cell(row=payload['row'] + 1, column=payload['column'] + 1)
                    cell.value = payload.get('value')

                elif op_type == BackendOperationType.IMPORT_DATAFRAME:
                    df = pd.DataFrame(payload['df']['data'], columns=payload['df']['columns'])
                    self._apply_dataframe_to_sheet(df, payload['sheet_name'], payload['start_row'],
                                                   payload['start_col'])

                elif op_type == BackendOperationType.REPLACE_SHEET_FROM_ANOTHER_WORKBOOK:
                    source_wb = openpyxl.load_workbook(BytesIO(payload['source_workbook_bytes']))
                    source_sheet = source_wb[payload['sheet_name']]
                    self._copy_sheet_between_workbooks(source_sheet, payload['new_sheet_name'])

            except Exception as e:
                self.logger.error(f"Failed to apply operation {op_type.name}: {e}", exc_info=True)

    def _apply_dataframe_to_sheet(self, df: pd.DataFrame, sheet_name: str, start_row: int, start_col: int):
        """Helper to write a DataFrame, preserving styles of untouched cells."""
        if sheet_name not in self._user_workbook.sheetnames:
            self._user_workbook.create_sheet(title=sheet_name)
        sheet = self._user_workbook[sheet_name]

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, start_col):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    def _copy_sheet_between_workbooks(self, source_sheet: Worksheet, new_sheet_name: str):
        """Helper for high-fidelity sheet copy."""
        if new_sheet_name in self._user_workbook.sheetnames:
            del self._user_workbook[new_sheet_name]
        target_sheet = self._user_workbook.create_sheet(title=new_sheet_name)

        # Copy data and styles cell by cell
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = copy.copy(cell.number_format)
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)

        # Copy other sheet properties
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))
        for key, dim in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[key] = copy.copy(dim)
        for key, dim in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[key] = copy.copy(dim)

    def apply_component_update(self, update_data: Dict[str, Any]):
        """Applies a single cell update from the UniverJS frontend."""
        if not self.has_workbook(): return
        sheet_name = update_data.get('sheet')
        values = update_data.get('value')
        if not sheet_name or not values or sheet_name not in self.workbook.sheetnames: return
        sheet = self.workbook[sheet_name]
        for row_idx, cols in values.items():
            for col_idx, cell_data in cols.items():
                cell = sheet.cell(row=int(row_idx) + 1, column=int(col_idx) + 1)
                if not cell.data_type == 'f':  # Preserve formulas
                    cell.value = cell_data.get('v')

    def save_workbook_to_bytes(self, workbook: openpyxl.Workbook) -> bytes:
        """
        Saves a workbook to an in-memory byte stream.
        """
        try:
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            self.logger.error(f"Erreur sauvegarde workbook: {str(e)}")
            try:
                self.logger.warning("Tentative de sauvegarde sans images...")
                return self._save_workbook_without_images(workbook)
            except:
                raise e

    @staticmethod
    def _save_workbook_without_images(workbook: openpyxl.Workbook) -> bytes:
        """Fallback to save workbook without images if standard save fails."""
        output = BytesIO()
        for sheet in workbook.worksheets:
            if hasattr(sheet, '_images'):
                sheet._images = []
        workbook.save(output)
        output.seek(0)
        UpdatedExcelHandler.logger.warning("Workbook sauvegardé sans images")
        return output.getvalue()