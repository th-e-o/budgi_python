import pandas as pd
import openpyxl
from typing import Optional, Dict, Any
import logging

from backend.core.excel_handler.excel_update_builder import ExcelUpdateBuilder

logger = logging.getLogger(__name__)


class BPSSTool:
    """
    A tool for processing BPSS budget files and registering the required
    workbook modifications into an ExcelUpdateBuilder instance.
    """

    def __init__(self):
        """Initializes the BPSSTool with default configuration values."""
        self.config = {
            'default_year': 2025,
            'default_ministry': '38',
            'default_program': '150'
        }

    def process_files(self, ppes_path: str, dpp18_path: str, bud45_path: str,
                      year: int, ministry_code: str, program_code: str,
                      target_workbook: openpyxl.Workbook,
                      builder: ExcelUpdateBuilder) -> None:
        """
        Main processing function that loads source files and orchestrates
        the workbook modification tasks via the provided builder.
        The target_workbook is used for read-only checks (e.g., sheet existence).
        """
        try:
            # File existence checks (original logic)
            import os
            for path, name in [(ppes_path, 'PP-E-S'), (dpp18_path, 'DPP18'), (bud45_path, 'BUD45')]:
                if not os.path.exists(path):
                    raise FileNotFoundError(f"Fichier {name} introuvable: {path}")
                logger.info(f"Fichier {name} trouvé: {path}")

            # Load data from source Excel files (original logic)
            logger.info("Chargement des fichiers BPSS...")
            sheet_names = self._get_sheet_names(ministry_code, program_code)
            try:
                df_pp_categ = pd.read_excel(ppes_path, sheet_name=sheet_names['pp_categ'])
                df_entrants = pd.read_excel(ppes_path, sheet_name=sheet_names['entrants'])
                df_sortants = pd.read_excel(ppes_path, sheet_name=sheet_names['sortants'])
            except Exception as e:
                logger.warning(f"Erreur avec les noms de feuilles calculés: {str(e)}")
                try:
                    xl_file = pd.ExcelFile(ppes_path)
                    available_sheets = xl_file.sheet_names
                    logger.info(f"Feuilles disponibles dans PP-E-S: {available_sheets}")
                    pp_categ_sheet = entrants_sheet = sortants_sheet = None
                    for sheet in available_sheets:
                        sheet_lower = sheet.lower()
                        if 'pp_categ' in sheet_lower or 'categ' in sheet_lower: pp_categ_sheet = sheet
                        elif 'entrant' in sheet_lower: entrants_sheet = sheet
                        elif 'sortant' in sheet_lower: sortants_sheet = sheet
                    if not all([pp_categ_sheet, entrants_sheet, sortants_sheet]):
                        if len(available_sheets) >= 3:
                            pp_categ_sheet = pp_categ_sheet or available_sheets[0]
                            entrants_sheet = entrants_sheet or available_sheets[1]
                            sortants_sheet = sortants_sheet or available_sheets[2]
                        else: raise ValueError(f"Le fichier PP-E-S doit contenir au moins 3 feuilles, trouvé: {len(available_sheets)}")
                    logger.info(f"Utilisation des feuilles: {pp_categ_sheet}, {entrants_sheet}, {sortants_sheet}")
                    df_pp_categ = pd.read_excel(ppes_path, sheet_name=pp_categ_sheet)
                    df_entrants = pd.read_excel(ppes_path, sheet_name=entrants_sheet)
                    df_sortants = pd.read_excel(ppes_path, sheet_name=sortants_sheet)
                except Exception as e2:
                    logger.error(f"Impossible de charger PP-E-S: {str(e2)}")
                    raise
            try:
                df_dpp18 = pd.read_excel(dpp18_path)
            except Exception as e:
                logger.error(f"Erreur chargement DPP18: {str(e)}")
                df_dpp18 = pd.read_excel(dpp18_path, sheet_name=0)
            try:
                df_bud45 = pd.read_excel(bud45_path)
            except Exception as e:
                logger.error(f"Erreur chargement BUD45: {str(e)}")
                df_bud45 = pd.read_excel(bud45_path, sheet_name=0)

            # Apply treatments by registering operations with the builder
            self._completion_accueil(
                target_workbook, builder, year, ministry_code, program_code
            )
            self._load_ppes_data(
                target_workbook, builder, df_pp_categ, df_entrants, df_sortants, program_code
            )
            self._load_dpp18_data(
                target_workbook, builder, df_dpp18, program_code
            )
            self._load_bud45_data(
                target_workbook, builder, df_bud45, program_code
            )

            logger.info("Traitement BPSS terminé avec succès. Opérations enregistrées dans le builder.")

        except Exception as e:
            logger.error(f"Erreur traitement BPSS: {str(e)}")
            raise

    def _get_sheet_names(self, ministry_code: str, program_code: str) -> Dict[str, str]:
        """Generates the expected sheet names based on ministry and program codes."""
        return {
            'pp_categ': f"MIN_{ministry_code}_DETAIL_Prog_PP_CATEG",
            'entrants': f"MIN_{ministry_code}_DETAIL_Prog_Entrants",
            'sortants': f"MIN_{ministry_code}_DETAIL_Prog_Sortants"
        }

    def _completion_accueil(self, wb: openpyxl.Workbook, builder: ExcelUpdateBuilder, year: int,
                            ministry_code: str, program_code: str) -> None:
        """Fills metadata in the 'Accueil' sheet (year, ministry, program)."""
        sheet_name = 'Accueil'
        # Note: builder uses 0-based indexing, openpyxl uses 1-based.
        builder.update_cell_value(sheet_name, row=34 - 1, col=3 - 1, value=year)
        builder.update_cell_value(sheet_name, row=35 - 1, col=3 - 1, value=year + 1)
        builder.update_cell_value(sheet_name, row=37 - 1, col=4 - 1, value=ministry_code)
        builder.update_cell_value(sheet_name, row=39 - 1, col=4 - 1, value=program_code)

    def _load_ppes_data(self, wb: openpyxl.Workbook, builder: ExcelUpdateBuilder,
                        df_pp_categ: pd.DataFrame, df_entrants: pd.DataFrame,
                        df_sortants: pd.DataFrame, program_code: str) -> None:
        """
        Loads data from the three PP-E-S dataframes into the 'Données PP-E-S'
        sheet and updates the 'Accueil' sheet with 'Indicié' data.
        """
        code_prefix = program_code[:3]
        df1 = df_pp_categ[df_pp_categ['nom_prog'].astype(str).str[:3] == code_prefix]
        df2 = df_entrants[df_entrants['nom_prog'].astype(str).str[:3] == code_prefix]
        df3 = df_sortants[df_sortants['nom_prog'].astype(str).str[:3] == code_prefix]

        sheet_name = "Données PP-E-S"
        if sheet_name not in wb.sheetnames:
            builder.create_sheet(sheet_name)

        start_rows = [7, 106, 205]
        dataframes = [df1, df2, df3]
        for start_row, df in zip(start_rows, dataframes):
            df_limited = df.head(100)
            for r_idx, row in enumerate(df_limited.values):
                for c_idx, value in enumerate(row):
                    builder.update_cell_value(sheet_name, row=start_row + r_idx - 1, col=3 + c_idx - 1, value=value)
                if start_row == 7:
                    builder.update_cell_value(sheet_name, row=start_row + r_idx - 1, col=2 - 1, value=row[3][:4])
                else:
                    builder.update_cell_value(sheet_name, row=start_row + r_idx - 1, col=2 - 1, value=row[2][:4])

        # Special handling for 'Indicié' data
        df_indicie = None
        for col in df1.columns:
            if 'Indicié' in df1[col].astype(str).values:
                df_indicie = df1[df1[col] == 'Indicié']
                logger.info(f"Colonne 'Indicié' trouvée : {col}")
                break
        if df_indicie is None or df_indicie.empty:
            logger.warning("Aucune donnée 'Indicié' trouvée dans PP-E-S")
        else:
            accueil_sheet_name = 'Accueil'
            if accueil_sheet_name not in wb.sheetnames:
                builder.create_sheet(accueil_sheet_name)
            # Clear previous data in range B43:C53.
            for row in range(43, 54):
                builder.update_cell_value(accueil_sheet_name, row=row - 1, col=2 - 1, value=None)
                builder.update_cell_value(accueil_sheet_name, row=row - 1, col=3 - 1, value=None)
            # Write new data
            row_dest = 43
            for idx, row_data in enumerate(df_indicie.values):
                if row_dest > 54: break
                code_categorie = row_data[3][:4]
                nom_categorie = row_data[3]
                if code_categorie and nom_categorie:
                    builder.update_cell_value(accueil_sheet_name, row=row_dest - 1, col=2 - 1, value=code_categorie)
                    builder.update_cell_value(accueil_sheet_name, row=row_dest - 1, col=3 - 1, value=nom_categorie)
                    row_dest += 1
                else:
                    logger.warning(f"Impossible d'extraire code/nom de la ligne: {row_data[:5]}")
        logger.info("Opérations de chargement PP-E-S enregistrées.")

    def _load_dpp18_data(self, wb: openpyxl.Workbook, builder: ExcelUpdateBuilder,
                         df_dpp18: pd.DataFrame, program_code: str) -> None:
        """Loads data from the DPP18 dataframe into the 'INF DPP 18' sheet."""
        sheet_name = "INF DPP 18"
        if sheet_name not in wb.sheetnames:
            builder.create_sheet(sheet_name)

        # Write header (first 5 rows)
        header = df_dpp18.head(5)
        for r_idx, row in enumerate(header.values):
            for c_idx, value in enumerate(row):
                builder.update_cell_value(sheet_name, row=2 + r_idx - 1, col=1 + c_idx - 1, value=value)

        code_prefix = program_code[:3]
        df_filtered = df_dpp18[df_dpp18.iloc[:, 0].astype(str).str.contains(code_prefix, na=False)]

        # Write filtered data, replicating the original's specific logic.
        for r_idx, row in enumerate(df_filtered.values):
            # First, write the entire row starting at row 7.
            for c_idx, value in enumerate(row):
                builder.update_cell_value(sheet_name, row=7 + r_idx - 1, col=1 + c_idx - 1, value=value)
            # Then, write a specific calculated value to column A, starting at row 6.
            builder.update_cell_value(sheet_name, row=6 + r_idx - 1, col=1 - 1, value=row[1][:14])
        logger.info("Opérations de chargement DPP18 enregistrées.")

    def _load_bud45_data(self, wb: openpyxl.Workbook, builder: ExcelUpdateBuilder,
                         df_bud45: pd.DataFrame, program_code: str) -> None:
        """Loads data from the BUD45 dataframe into the 'INF BUD 45' sheet."""
        sheet_name = "INF BUD 45"
        if sheet_name not in wb.sheetnames:
            builder.create_sheet(sheet_name)

        header = df_bud45.head(5)
        for r_idx, row in enumerate(header.values):
            for c_idx, value in enumerate(row):
                builder.update_cell_value(sheet_name, row=2 + r_idx - 1, col=1 + c_idx - 1, value=value)

        if len(df_bud45.columns) > 1:
            code_prefix = program_code[:3]
            df_filtered = df_bud45[df_bud45.iloc[:, 1].astype(str).str.contains(code_prefix, na=False)]
            # Write filtered data, replicating the write-then-overwrite logic.
            for r_idx, row in enumerate(df_filtered.values):
                # First, write the entire row.
                for c_idx, value in enumerate(row):
                    builder.update_cell_value(sheet_name, row=7 + r_idx - 1, col=1 + c_idx - 1, value=value)
                # Immediately overwrite the value in the first column of that same row.
                builder.update_cell_value(sheet_name, row=7 + r_idx - 1, col=1 - 1, value=row[3][:14])
        logger.info("Opérations de chargement BUD45 enregistrées.")